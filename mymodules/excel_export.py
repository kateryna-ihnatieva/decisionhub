import io
from datetime import datetime
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from mymodules.mai import generate_hierarchy_tree


class HierarchyExcelExporter:
    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def create_workbook(self):
        """Create a new Excel workbook"""
        self.workbook = openpyxl.Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)

    def add_worksheet(self, title):
        """Add a new worksheet to the workbook"""
        self.worksheet = self.workbook.create_sheet(title=title)
        return self.worksheet

    def set_header_style(self, cell, text, color="366092"):
        """Set header cell style"""
        cell.value = text
        cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_subheader_style(self, cell, text, color="4F81BD"):
        """Set subheader cell style"""
        cell.value = text
        cell.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_data_style(self, cell, value, bold=False):
        """Set data cell style"""
        cell.value = value
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_number_style(self, cell, value, decimal_places=4, bold=False):
        """Set number cell style with specific decimal places"""
        if isinstance(value, (int, float)):
            cell.value = round(value, decimal_places)
        elif isinstance(value, str) and "/" in value:
            # Convert string fractions to float
            try:
                numerator, denominator = value.split("/")
                cell.value = round(
                    float(numerator) / float(denominator), decimal_places
                )
            except (ValueError, ZeroDivisionError):
                cell.value = 0.0
        else:
            cell.value = value
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_percentage_style(self, cell, value, decimal_places=3, bold=False):
        """Set percentage cell style - values are already in percentage format"""
        if isinstance(value, (int, float)):
            cell.value = value  # Values are already in percentage format from database
        else:
            cell.value = 0.0
        cell.number_format = f"0.{'0' * decimal_places}%"
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def convert_fraction_to_float(self, value):
        """Convert string fraction to float"""
        if isinstance(value, (int, float)):
            return float(value)
        elif isinstance(value, str) and "/" in value:
            try:
                numerator, denominator = value.split("/")
                return float(numerator) / float(denominator)
            except (ValueError, ZeroDivisionError):
                return 0.0
        else:
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0

    def auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths with better formatting for wrapped text"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        # Count characters, considering text wrapping
                        cell_length = len(str(cell.value))
                        if cell.alignment and cell.alignment.wrap_text:
                            # For wrapped text, consider line breaks
                            lines = str(cell.value).count("\n") + 1
                            cell_length = max(
                                len(line) for line in str(cell.value).split("\n")
                            )
                        if cell_length > max_length:
                            max_length = cell_length
                except (ValueError, TypeError):
                    pass
            # Better width calculation for wrapped text
            adjusted_width = min(max(max_length + 2, 15), 60)  # Min 15, max 60
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def create_general_info_sheet(self, task_description, method_id, created_date=None):
        """Create general information sheet"""
        ws = self.add_worksheet("Загальна інформація")

        # Title
        self.set_header_style(ws["A1"], "Звіт аналізу AHP", "366092")
        ws.merge_cells("A1:D1")

        # Method information
        ws["A3"] = "Метод аналізу:"
        ws["B3"] = "Метод аналізу ієрархій (AHP)"
        ws["A4"] = "ID аналізу:"
        ws["B4"] = f"Завдання_{method_id}"
        ws["A5"] = "Дата створення:"
        ws["B5"] = created_date or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Task description
        if task_description:
            ws["A7"] = "Опис завдання:"
            ws["A8"] = task_description
            ws.merge_cells("A8:D12")
            ws["A8"].alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

        # Style the information cells
        for row in range(3, 6):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name="Arial", size=11, bold=(col == 1))
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        self.auto_adjust_columns(ws)

    def create_criteria_sheet(self, criteria_names, criteria_weights):
        """Create criteria sheet with weights and ranking"""
        ws = self.add_worksheet("Критерії")

        # Headers
        self.set_header_style(ws["A1"], "Аналіз критеріїв", "366092")
        ws.merge_cells("A1:D1")

        self.set_subheader_style(ws["A3"], "Критерій", "4F81BD")
        self.set_subheader_style(ws["B3"], "Вага", "4F81BD")
        self.set_subheader_style(ws["C3"], "Відсоток", "4F81BD")

        # Data
        for i, (name, weight) in enumerate(zip(criteria_names, criteria_weights)):
            row = i + 4
            self.set_data_style(ws.cell(row=row, column=1), name)
            self.set_number_style(ws.cell(row=row, column=2), weight)
            self.set_number_style(ws.cell(row=row, column=3), weight * 100, 2)
            ws.cell(row=row, column=3).number_format = "0.00%"

        self.auto_adjust_columns(ws)

    def create_alternatives_sheet(self, alternatives_names, global_priorities):
        """Create alternatives sheet with global priorities"""
        ws = self.add_worksheet("Альтернативи")

        # Headers
        self.set_header_style(ws["A1"], "Аналіз альтернатив", "366092")
        ws.merge_cells("A1:D1")

        self.set_subheader_style(ws["A3"], "Альтернатива", "4F81BD")
        self.set_subheader_style(ws["B3"], "Глобальний пріоритет", "4F81BD")
        self.set_subheader_style(ws["C3"], "Відсоток", "4F81BD")

        # Data
        for i, (name, priority) in enumerate(
            zip(alternatives_names, global_priorities)
        ):
            row = i + 4
            self.set_data_style(ws.cell(row=row, column=1), name)
            self.set_number_style(ws.cell(row=row, column=2), priority)
            self.set_number_style(ws.cell(row=row, column=3), priority * 100, 2)
            ws.cell(row=row, column=3).number_format = "0.00%"

        self.auto_adjust_columns(ws)

    def create_criteria_matrix_sheet(
        self, criteria_names, matrix_data, eigenvector, normalized_eigenvector
    ):
        """Create criteria comparison matrix sheet with additional calculations"""
        ws = self.add_worksheet("Матриця критеріїв")

        # Headers
        self.set_header_style(ws["A1"], "Матриця порівняння критеріїв", "366092")
        ws.merge_cells(f"A1:{get_column_letter(len(criteria_names) + 3)}1")

        # Matrix headers
        self.set_subheader_style(ws["A3"], "", "4F81BD")
        for i, name in enumerate(criteria_names):
            col = i + 2
            self.set_subheader_style(ws.cell(row=3, column=col), name, "4F81BD")
        self.set_subheader_style(
            ws.cell(row=3, column=len(criteria_names) + 2), "Власний вектор", "4F81BD"
        )
        self.set_subheader_style(
            ws.cell(row=3, column=len(criteria_names) + 3), "Нормалізований", "4F81BD"
        )

        # Matrix data
        for i, name in enumerate(criteria_names):
            row = i + 4
            self.set_subheader_style(ws.cell(row=row, column=1), name, "4F81BD")
            for j in range(len(criteria_names)):
                col = j + 2
                self.set_number_style(ws.cell(row=row, column=col), matrix_data[i][j])
            self.set_number_style(
                ws.cell(row=row, column=len(criteria_names) + 2), eigenvector[i]
            )
            self.set_number_style(
                ws.cell(row=row, column=len(criteria_names) + 3),
                normalized_eigenvector[i],
            )

        # Sum row (only for last two columns)
        sum_row = len(criteria_names) + 4
        self.set_subheader_style(ws.cell(row=sum_row, column=1), "Сума", "4F81BD")
        # Empty cells for matrix columns
        for j in range(len(criteria_names)):
            col = j + 2
            self.set_data_style(ws.cell(row=sum_row, column=col), "")
        # Sum of eigenvector
        eigenvector_sum = sum(eigenvector)
        self.set_number_style(
            ws.cell(row=sum_row, column=len(criteria_names) + 2), eigenvector_sum
        )
        # Sum of normalized (should be 1)
        normalized_sum = sum(normalized_eigenvector)
        self.set_number_style(
            ws.cell(row=sum_row, column=len(criteria_names) + 3), normalized_sum
        )

        # Additional calculations section
        calc_start_row = sum_row + 2

        # Sum by columns row
        self.set_subheader_style(
            ws.cell(row=calc_start_row, column=1), "Сума по стовпцях", "4F81BD"
        )
        for j in range(len(criteria_names)):
            col = j + 2
            col_sum = sum(
                self.convert_fraction_to_float(matrix_data[i][j])
                for i in range(len(criteria_names))
            )
            self.set_number_style(ws.cell(row=calc_start_row, column=col), col_sum)
        self.set_subheader_style(
            ws.cell(row=calc_start_row, column=len(criteria_names) + 2),
            "Всього (Lmax)",
            "4F81BD",
        )

        # Product row
        product_row = calc_start_row + 1
        self.set_subheader_style(
            ws.cell(row=product_row, column=1),
            "Добуток додатку по стовпцях і нормалізованої оцінки вектора пріоритету",
            "4F81BD",
        )
        for j in range(len(criteria_names)):
            col = j + 2
            col_sum = sum(
                self.convert_fraction_to_float(matrix_data[i][j])
                for i in range(len(criteria_names))
            )
            product = col_sum * normalized_eigenvector[j]
            self.set_number_style(ws.cell(row=product_row, column=col), product)
        # Lmax calculation
        l_max = sum(
            sum(
                self.convert_fraction_to_float(matrix_data[i][j])
                for i in range(len(criteria_names))
            )
            * normalized_eigenvector[j]
            for j in range(len(criteria_names))
        )
        self.set_number_style(
            ws.cell(row=product_row, column=len(criteria_names) + 2), l_max
        )

        # Consistency indicators
        consistency_start_row = product_row + 3
        self.set_header_style(
            ws.cell(row=consistency_start_row, column=1),
            "Показники узгодженості",
            "366092",
        )
        ws.merge_cells(f"A{consistency_start_row}:D{consistency_start_row}")

        # Calculate consistency
        n = len(criteria_names)
        ci = (l_max - n) / (n - 1) if n > 1 else 0
        ri_values = {
            1: 0,
            2: 0,
            3: 0.58,
            4: 0.9,
            5: 1.12,
            6: 1.24,
            7: 1.32,
            8: 1.41,
            9: 1.45,
            10: 1.49,
        }
        ri = ri_values.get(n, 1.49)
        cr = ci / ri if ri > 0 else 0

        self.set_subheader_style(
            ws.cell(row=consistency_start_row + 2, column=1),
            "Індекс узгодженості (ІУ)",
            "4F81BD",
        )
        self.set_number_style(ws.cell(row=consistency_start_row + 2, column=2), ci)

        self.set_subheader_style(
            ws.cell(row=consistency_start_row + 3, column=1),
            "Відношення узгодженості (ВУ)",
            "4F81BD",
        )
        self.set_percentage_style(ws.cell(row=consistency_start_row + 3, column=2), cr)

        self.set_subheader_style(
            ws.cell(row=consistency_start_row + 4, column=1), "Статус", "4F81BD"
        )
        status = "Прийнятно" if cr <= 0.1 else "Потребує перегляду"
        self.set_data_style(ws.cell(row=consistency_start_row + 4, column=2), status)

        # Ranking
        ranking_start_row = consistency_start_row + 6
        self.set_header_style(
            ws.cell(row=ranking_start_row, column=1), "Ранжування критеріїв", "366092"
        )
        ws.merge_cells(f"A{ranking_start_row}:D{ranking_start_row}")

        # Create ranking text
        ranking_pairs = list(zip(criteria_names, normalized_eigenvector))
        ranking_pairs.sort(key=lambda x: x[1], reverse=True)
        ranking_text = " > ".join([name for name, _ in ranking_pairs])
        self.set_data_style(ws.cell(row=ranking_start_row + 2, column=1), ranking_text)
        ws.merge_cells(f"A{ranking_start_row + 2}:D{ranking_start_row + 2}")

        self.auto_adjust_columns(ws)

    def create_alternatives_matrix_sheet(
        self,
        criteria_name,
        alternatives_names,
        matrix_data,
        eigenvector,
        normalized_eigenvector,
    ):
        """Create alternatives comparison matrix sheet for a specific criteria with additional calculations"""
        ws = self.add_worksheet(f"Матриця_{criteria_name[:20]}")

        # Headers
        self.set_header_style(
            ws["A1"], f"Матриця альтернатив - {criteria_name}", "366092"
        )
        ws.merge_cells(f"A1:{get_column_letter(len(alternatives_names) + 3)}1")

        # Matrix headers
        self.set_subheader_style(ws["A3"], "", "4F81BD")
        for i, name in enumerate(alternatives_names):
            col = i + 2
            self.set_subheader_style(ws.cell(row=3, column=col), name, "4F81BD")
        self.set_subheader_style(
            ws.cell(row=3, column=len(alternatives_names) + 2),
            "Власний вектор",
            "4F81BD",
        )
        self.set_subheader_style(
            ws.cell(row=3, column=len(alternatives_names) + 3),
            "Нормалізований",
            "4F81BD",
        )

        # Matrix data
        for i, name in enumerate(alternatives_names):
            row = i + 4
            self.set_subheader_style(ws.cell(row=row, column=1), name, "4F81BD")
            for j in range(len(alternatives_names)):
                col = j + 2
                self.set_number_style(ws.cell(row=row, column=col), matrix_data[i][j])
            self.set_number_style(
                ws.cell(row=row, column=len(alternatives_names) + 2), eigenvector[i]
            )
            self.set_number_style(
                ws.cell(row=row, column=len(alternatives_names) + 3),
                normalized_eigenvector[i],
            )

        # Sum row (only for last two columns)
        sum_row = len(alternatives_names) + 4
        self.set_subheader_style(ws.cell(row=sum_row, column=1), "Сума", "4F81BD")
        # Empty cells for matrix columns
        for j in range(len(alternatives_names)):
            col = j + 2
            self.set_data_style(ws.cell(row=sum_row, column=col), "")
        # Sum of eigenvector
        eigenvector_sum = sum(eigenvector)
        self.set_number_style(
            ws.cell(row=sum_row, column=len(alternatives_names) + 2), eigenvector_sum
        )
        # Sum of normalized (should be 1)
        normalized_sum = sum(normalized_eigenvector)
        self.set_number_style(
            ws.cell(row=sum_row, column=len(alternatives_names) + 3), normalized_sum
        )

        # Additional calculations section
        calc_start_row = sum_row + 2

        # Sum by columns row
        self.set_subheader_style(
            ws.cell(row=calc_start_row, column=1), "Сума по стовпцях", "4F81BD"
        )
        for j in range(len(alternatives_names)):
            col = j + 2
            col_sum = sum(
                self.convert_fraction_to_float(matrix_data[i][j])
                for i in range(len(alternatives_names))
            )
            self.set_number_style(ws.cell(row=calc_start_row, column=col), col_sum)
        self.set_subheader_style(
            ws.cell(row=calc_start_row, column=len(alternatives_names) + 2),
            "Всього (Lmax)",
            "4F81BD",
        )

        # Product row
        product_row = calc_start_row + 1
        self.set_subheader_style(
            ws.cell(row=product_row, column=1),
            "Добуток додатку по стовпцях і нормалізованої оцінки вектора пріоритету",
            "4F81BD",
        )
        for j in range(len(alternatives_names)):
            col = j + 2
            col_sum = sum(
                self.convert_fraction_to_float(matrix_data[i][j])
                for i in range(len(alternatives_names))
            )
            product = col_sum * normalized_eigenvector[j]
            self.set_number_style(ws.cell(row=product_row, column=col), product)
        # Lmax calculation
        l_max = sum(
            sum(
                self.convert_fraction_to_float(matrix_data[i][j])
                for i in range(len(alternatives_names))
            )
            * normalized_eigenvector[j]
            for j in range(len(alternatives_names))
        )
        self.set_number_style(
            ws.cell(row=product_row, column=len(alternatives_names) + 2), l_max
        )

        # Consistency indicators
        consistency_start_row = product_row + 3
        self.set_header_style(
            ws.cell(row=consistency_start_row, column=1),
            "Показники узгодженості",
            "366092",
        )
        ws.merge_cells(f"A{consistency_start_row}:D{consistency_start_row}")

        # Calculate consistency
        n = len(alternatives_names)
        ci = (l_max - n) / (n - 1) if n > 1 else 0
        ri_values = {
            1: 0,
            2: 0,
            3: 0.58,
            4: 0.9,
            5: 1.12,
            6: 1.24,
            7: 1.32,
            8: 1.41,
            9: 1.45,
            10: 1.49,
        }
        ri = ri_values.get(n, 1.49)
        cr = ci / ri if ri > 0 else 0

        self.set_subheader_style(
            ws.cell(row=consistency_start_row + 2, column=1),
            "Індекс узгодженості (ІУ)",
            "4F81BD",
        )
        self.set_number_style(ws.cell(row=consistency_start_row + 2, column=2), ci)

        self.set_subheader_style(
            ws.cell(row=consistency_start_row + 3, column=1),
            "Відношення узгодженості (ВУ)",
            "4F81BD",
        )
        self.set_percentage_style(ws.cell(row=consistency_start_row + 3, column=2), cr)

        self.set_subheader_style(
            ws.cell(row=consistency_start_row + 4, column=1), "Статус", "4F81BD"
        )
        status = "Прийнятно" if cr < 10 else "Потребує перегляду"
        self.set_data_style(ws.cell(row=consistency_start_row + 4, column=2), status)

        # Ranking
        ranking_start_row = consistency_start_row + 6
        self.set_header_style(
            ws.cell(row=ranking_start_row, column=1),
            f"Ранжування альтернатив ({criteria_name})",
            "366092",
        )
        ws.merge_cells(f"A{ranking_start_row}:D{ranking_start_row}")

        # Create ranking text
        ranking_pairs = list(zip(alternatives_names, normalized_eigenvector))
        ranking_pairs.sort(key=lambda x: x[1], reverse=True)
        ranking_text = " > ".join([name for name, _ in ranking_pairs])
        self.set_data_style(ws.cell(row=ranking_start_row + 2, column=1), ranking_text)
        ws.merge_cells(f"A{ranking_start_row + 2}:D{ranking_start_row + 2}")

        self.auto_adjust_columns(ws)

    def create_results_sheet(
        self,
        alternatives_names,
        criteria_names,
        criteria_weights,
        alternatives_weights,
        global_priorities,
    ):
        """Create final results sheet"""
        ws = self.add_worksheet("Результати")

        # Headers
        self.set_header_style(ws["A1"], "Фінальні результати", "366092")
        ws.merge_cells(f"A1:{get_column_letter(len(criteria_names) + 3)}1")

        # Table headers
        self.set_subheader_style(ws["A3"], "Альтернативи", "4F81BD")
        for i, name in enumerate(criteria_names):
            col = i + 2
            self.set_subheader_style(ws.cell(row=3, column=col), name, "4F81BD")
        self.set_subheader_style(
            ws.cell(row=3, column=len(criteria_names) + 2),
            "Глобальний пріоритет",
            "4F81BD",
        )

        # Criteria weights row
        self.set_subheader_style(ws.cell(row=4, column=1), "Ваги критеріїв", "4F81BD")
        for i, weight in enumerate(criteria_weights):
            col = i + 2
            self.set_number_style(ws.cell(row=4, column=col), weight)

        # Alternatives data
        for i, name in enumerate(alternatives_names):
            row = i + 5
            self.set_data_style(ws.cell(row=row, column=1), name)
            for j in range(len(criteria_names)):
                col = j + 2
                self.set_number_style(
                    ws.cell(row=row, column=col), alternatives_weights[j][i]
                )
            self.set_number_style(
                ws.cell(row=row, column=len(criteria_names) + 2),
                global_priorities[i],
                bold=True,
            )

        # Sum row for global priorities
        sum_row = len(alternatives_names) + 5
        self.set_subheader_style(ws.cell(row=sum_row, column=1), "Сума", "4F81BD")
        # Empty cells for criteria columns
        for j in range(len(criteria_names)):
            col = j + 2
            self.set_data_style(ws.cell(row=sum_row, column=col), "")
        # Sum of global priorities (should be 1)
        global_sum = sum(global_priorities)
        self.set_number_style(
            ws.cell(row=sum_row, column=len(criteria_names) + 2), global_sum, bold=True
        )

        self.auto_adjust_columns(ws)

    def create_consistency_sheet(
        self, criteria_consistency, alternatives_consistency, criteria_names
    ):
        """Create consistency indicators sheet"""
        ws = self.add_worksheet("Узгодженість")

        # Headers
        self.set_header_style(ws["A1"], "Показники узгодженості", "366092")
        ws.merge_cells("A1:D1")

        # Criteria consistency
        self.set_subheader_style(ws["A3"], "Матриця критеріїв", "4F81BD")
        self.set_subheader_style(ws["B3"], "ІУ", "4F81BD")
        self.set_subheader_style(ws["C3"], "ВУ", "4F81BD")
        self.set_subheader_style(ws["D3"], "Статус", "4F81BD")

        self.set_data_style(ws.cell(row=4, column=1), "Критерії")
        self.set_number_style(ws.cell(row=4, column=2), criteria_consistency["ci"])
        self.set_percentage_style(ws.cell(row=4, column=3), criteria_consistency["cr"])
        status = (
            "Прийнятно" if criteria_consistency["cr"] <= 0.1 else "Потребує перегляду"
        )
        self.set_data_style(ws.cell(row=4, column=4), status)

        # Alternatives consistency
        row = 6
        self.set_subheader_style(
            ws.cell(row=row, column=1), "Матриці альтернатив", "4F81BD"
        )
        self.set_subheader_style(ws.cell(row=row, column=2), "ІУ", "4F81BD")
        self.set_subheader_style(ws.cell(row=row, column=3), "ВУ", "4F81BD")
        self.set_subheader_style(ws.cell(row=row, column=4), "Статус", "4F81BD")

        for i, (ci, cr) in enumerate(
            zip(alternatives_consistency["ci"], alternatives_consistency["cr"])
        ):
            row = i + 7
            self.set_data_style(ws.cell(row=row, column=1), criteria_names[i])
            self.set_number_style(ws.cell(row=row, column=2), ci)
            self.set_percentage_style(ws.cell(row=row, column=3), cr)
            status = "Acceptable" if cr <= 0.1 else "Review Required"
            self.set_data_style(ws.cell(row=row, column=4), status)

        self.auto_adjust_columns(ws)

    def create_chart_sheet(
        self,
        alternatives_names,
        global_priorities,
        criteria_names=None,
        criteria_weights=None,
    ):
        """Create chart sheet with visualization and hierarchical representation"""
        ws = self.add_worksheet("Графіки")

        # Headers
        self.set_header_style(ws["A1"], "Візуалізація результатів", "366092")
        ws.merge_cells("A1:H1")

        # Prepare data for chart
        ws["A3"] = "Альтернатива"
        ws["B3"] = "Глобальний пріоритет"

        for i, (name, priority) in enumerate(
            zip(alternatives_names, global_priorities)
        ):
            ws.cell(row=i + 4, column=1).value = name
            ws.cell(row=i + 4, column=2).value = priority

        # Create bar chart
        chart = BarChart()
        chart.title = "Порівняння глобальних пріоритетів"
        chart.x_axis.title = "Альтернативи"
        chart.y_axis.title = "Пріоритет"

        data = Reference(ws, min_col=2, min_row=3, max_row=len(alternatives_names) + 3)
        categories = Reference(
            ws, min_col=1, min_row=4, max_row=len(alternatives_names) + 3
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "E3")

        # Add hierarchical representation as image
        if criteria_names and criteria_weights:
            self.create_hierarchy_image(
                ws,
                criteria_names,
                alternatives_names,
                criteria_weights,
                global_priorities,
            )

        self.auto_adjust_columns(ws)

    def create_hierarchy_image(
        self,
        ws,
        criteria_names,
        alternatives_names,
        criteria_weights,
        global_priorities,
    ):
        """Create hierarchical representation using existing generate_hierarchy_tree function"""
        try:
            # Generate hierarchy tree using existing function
            generate_hierarchy_tree(
                criteria_names, alternatives_names, criteria_weights, global_priorities
            )

            # The function saves to static/img/hierarchy_tree.png, so we need to read it
            source_path = "static/img/hierarchy_tree.png"
            if os.path.exists(source_path):
                # Read the image file into BytesIO
                with open(source_path, "rb") as img_file:
                    img_data = io.BytesIO(img_file.read())

                # Add image to Excel with proper aspect ratio
                img = Image(img_data)

                # Calculate proper dimensions maintaining aspect ratio
                # Original image is typically wider, so we set a good width and calculate height
                target_width = 800  # Wider for better visibility
                aspect_ratio = img.height / img.width
                target_height = int(target_width * aspect_ratio)

                # Ensure reasonable bounds
                if target_height > 600:
                    target_height = 600
                    target_width = int(target_height / aspect_ratio)

                img.width = target_width
                img.height = target_height

                # Position image below the chart with more spacing
                ws.add_image(img, "A20")
            else:
                print(f"Hierarchy tree image not found at {source_path}")
                # Fallback to text representation
                self.create_hierarchical_diagram(
                    ws,
                    criteria_names,
                    alternatives_names,
                    criteria_weights,
                    global_priorities,
                )

        except Exception as e:
            print(f"Error creating hierarchy image: {e}")
            # Fallback to text representation
            self.create_hierarchical_diagram(
                ws,
                criteria_names,
                alternatives_names,
                criteria_weights,
                global_priorities,
            )

    def create_hierarchical_diagram(
        self,
        ws,
        criteria_names,
        alternatives_names,
        criteria_weights,
        global_priorities,
    ):
        """Create hierarchical representation diagram"""
        start_row = len(alternatives_names) + 8

        # Title for hierarchical representation
        self.set_subheader_style(
            ws.cell(row=start_row, column=1), "Ієрархічне представлення AHP", "4F81BD"
        )
        ws.merge_cells(f"A{start_row}:H{start_row}")

        # Level 1: Goal
        goal_row = start_row + 2
        ws.cell(row=goal_row, column=1).value = "Рівень 1: Мета"
        ws.cell(row=goal_row, column=1).font = Font(
            name="Arial", size=12, bold=True, color="FFFFFF"
        )
        ws.cell(row=goal_row, column=1).fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        ws.cell(row=goal_row, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.merge_cells(f"A{goal_row}:H{goal_row}")

        ws.cell(row=goal_row + 1, column=1).value = "Вибір найкращої альтернативи"
        ws.cell(row=goal_row + 1, column=1).font = Font(
            name="Arial", size=12, bold=True
        )
        ws.cell(row=goal_row + 1, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.merge_cells(f"A{goal_row + 1}:H{goal_row + 1}")

        # Level 2: Criteria
        criteria_start_row = goal_row + 3
        ws.cell(row=criteria_start_row, column=1).value = "Рівень 2: Критерії"
        ws.cell(row=criteria_start_row, column=1).font = Font(
            name="Arial", size=12, bold=True, color="FFFFFF"
        )
        ws.cell(row=criteria_start_row, column=1).fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        ws.cell(row=criteria_start_row, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.merge_cells(f"A{criteria_start_row}:H{criteria_start_row}")

        # Criteria with weights
        for i, (criteria, weight) in enumerate(zip(criteria_names, criteria_weights)):
            row = criteria_start_row + 1 + i
            ws.cell(row=row, column=1).value = f"• {criteria}"
            ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ws.cell(row=row, column=2).value = f"({weight:.1%})"
            ws.cell(row=row, column=2).font = Font(name="Arial", size=12)
            ws.cell(row=row, column=2).number_format = "0.0%"
            ws.cell(row=row, column=2).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Level 3: Alternatives
        alt_start_row = criteria_start_row + len(criteria_names) + 2
        ws.cell(row=alt_start_row, column=1).value = "Рівень 3: Альтернативи"
        ws.cell(row=alt_start_row, column=1).font = Font(
            name="Arial", size=12, bold=True, color="FFFFFF"
        )
        ws.cell(row=alt_start_row, column=1).fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        ws.cell(row=alt_start_row, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.merge_cells(f"A{alt_start_row}:H{alt_start_row}")

        # Alternatives with global priorities
        for i, (alternative, priority) in enumerate(
            zip(alternatives_names, global_priorities)
        ):
            row = alt_start_row + 1 + i
            ws.cell(row=row, column=1).value = f"• {alternative}"
            ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ws.cell(row=row, column=2).value = f"({priority:.1%})"
            ws.cell(row=row, column=2).font = Font(name="Arial", size=12)
            ws.cell(row=row, column=2).number_format = "0.0%"
            ws.cell(row=row, column=2).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Add visual hierarchy with arrows
        arrow_start_row = alt_start_row + len(alternatives_names) + 3
        ws.cell(row=arrow_start_row, column=1).value = "↓"
        ws.cell(row=arrow_start_row, column=1).font = Font(
            name="Arial", size=14, bold=True
        )
        ws.cell(row=arrow_start_row, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        ws.cell(row=arrow_start_row + 1, column=1).value = "↓"
        ws.cell(row=arrow_start_row + 1, column=1).font = Font(
            name="Arial", size=14, bold=True
        )
        ws.cell(row=arrow_start_row + 1, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Add summary table
        summary_start_row = arrow_start_row + 4
        ws.cell(row=summary_start_row, column=1).value = (
            "Підсумкова таблиця пріоритетів"
        )
        ws.cell(row=summary_start_row, column=1).font = Font(
            name="Arial", size=14, bold=True
        )
        ws.cell(row=summary_start_row, column=1).fill = PatternFill(
            start_color="FFC000", end_color="FFC000", fill_type="solid"
        )
        ws.cell(row=summary_start_row, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.merge_cells(f"A{summary_start_row}:H{summary_start_row}")

        # Table headers
        headers = ["Альтернатива", "Глобальний пріоритет", "Відсоток"]
        for i, header in enumerate(headers):
            cell = ws.cell(row=summary_start_row + 2, column=i + 1)
            cell.value = header
            cell.font = Font(name="Arial", size=12, bold=True)
            cell.fill = PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Table data
        for i, (alternative, priority) in enumerate(
            zip(alternatives_names, global_priorities)
        ):
            row = summary_start_row + 3 + i
            ws.cell(row=row, column=1).value = alternative
            ws.cell(row=row, column=2).value = priority
            ws.cell(row=row, column=3).value = priority * 100

            # Format cells
            for col in range(3):
                cell = ws.cell(row=row, column=col + 1)
                cell.font = Font(name="Arial", size=12)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

            # Format numbers
            ws.cell(row=row, column=2).number_format = "0.000"
            ws.cell(row=row, column=3).number_format = "0.0%"

    def generate_hierarchy_analysis_excel(self, analysis_data):
        """Generate complete Excel file for hierarchy analysis

        Args:
            analysis_data: Dictionary containing all analysis data
        """
        self.create_workbook()

        # Extract data
        method_id = analysis_data["method_id"]
        task_description = analysis_data.get("task_description")
        criteria_names = analysis_data["criteria_names"]
        alternatives_names = analysis_data["alternatives_names"]
        criteria_weights = analysis_data["criteria_weights"]
        global_priorities = analysis_data["global_priorities"]
        criteria_matrix = analysis_data["criteria_matrix"]
        alternatives_matrices = analysis_data["alternatives_matrices"]
        criteria_eigenvector = analysis_data["criteria_eigenvector"]
        alternatives_eigenvectors = analysis_data["alternatives_eigenvectors"]
        alternatives_weights = analysis_data["alternatives_weights"]
        criteria_consistency = analysis_data["criteria_consistency"]
        alternatives_consistency = analysis_data["alternatives_consistency"]

        # Create sheets in specified order
        # 1. General Information (overview)
        self.create_general_info_sheet(task_description, method_id)

        # 2. Criteria Matrix (with additional calculations)
        self.create_criteria_matrix_sheet(
            criteria_names, criteria_matrix, criteria_eigenvector, criteria_weights
        )

        # 3. All Alternatives Matrices (for each criterion)
        for i, criteria_name in enumerate(criteria_names):
            self.create_alternatives_matrix_sheet(
                criteria_name,
                alternatives_names,
                alternatives_matrices[i],
                alternatives_eigenvectors[i],
                alternatives_weights[i],  # Use correct normalized weights
            )

        # 4. Results (summary)
        self.create_results_sheet(
            alternatives_names,
            criteria_names,
            criteria_weights,
            alternatives_weights,
            global_priorities,
        )

        # 5. Charts and visualization (appendix)
        self.create_chart_sheet(
            alternatives_names, global_priorities, criteria_names, criteria_weights
        )

        return self.workbook

    def save_to_bytes(self):
        """Save workbook to bytes for download"""
        if not self.workbook:
            raise ValueError("No workbook to save")

        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
