import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from datetime import datetime


class ExpertsExcelExporter:
    def __init__(self):
        self.workbook = Workbook()
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"  # Dark blue
        )
        self.subheader_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"  # Medium blue
        )
        self.data_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"  # Light gray
        )
        self.sum_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"  # Light blue
        )

        self.header_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        self.subheader_font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        self.data_font = Font(name="Arial", size=11, color="000000")
        self.sum_font = Font(name="Arial", size=11, bold=True, color="000000")

        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_header_style(self, cell):
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.border = self.thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_subheader_style(self, cell):
        cell.font = self.subheader_font
        cell.fill = self.subheader_fill
        cell.border = self.thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_data_style(self, cell):
        cell.font = self.data_font
        cell.fill = self.data_fill
        cell.border = self.thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_sum_style(self, cell):
        cell.font = self.sum_font
        cell.fill = self.sum_fill
        cell.border = self.thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_number_style(self, cell, value):
        if isinstance(value, (int, float)):
            cell.value = int(value) if value == int(value) else round(value, 3)
        else:
            cell.value = value
        cell.number_format = "0.000" if isinstance(value, float) else "0"
        self.set_data_style(cell)

    def set_row_height_for_text(self, ws, row, text, min_height=30):
        """Set row height based on text length for better wrapping"""
        if text and len(str(text)) > 50:  # For long text
            # Calculate approximate height needed
            text_length = len(str(text))
            lines_needed = max(1, text_length // 50)  # Approximate lines
            height = max(min_height, lines_needed * 15)  # 15 points per line
            ws.row_dimensions[row].height = height

    def create_general_info_sheet(self, analysis_data):
        ws = self.workbook.active
        ws.title = "Загальна інформація"

        # Header
        ws["A1"] = "Звіт аналізу Експертних Оцінок"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:D1")

        # Method info
        ws["A3"] = "Метод аналізу:"
        ws["B3"] = "Експертні Оцінки"
        self.set_data_style(ws["A3"])
        self.set_data_style(ws["B3"])

        ws["A4"] = "ID аналізу:"
        ws["B4"] = analysis_data.get("method_id", "N/A")
        self.set_data_style(ws["A4"])
        self.set_data_style(ws["B4"])

        ws["A5"] = "Дата створення:"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.set_data_style(ws["A5"])
        self.set_data_style(ws["B5"])

        ws["A6"] = "Опис завдання:"
        self.set_data_style(ws["A6"])

        task_description = analysis_data.get("experts_task", "Аналіз експертних оцінок")
        ws["A7"] = task_description
        self.set_data_style(ws["A7"])
        ws.merge_cells("A7:D7")
        self.set_row_height_for_text(ws, 7, task_description)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_competency_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Компетентність експертів")

        # Header
        ws["A1"] = "Результати оцінки компетентності експертів"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:E1")

        table_competency = analysis_data.get("table_competency", [])
        name_arguments = analysis_data.get("name_arguments", [])
        k_k = analysis_data.get("k_k", [])
        k_a = analysis_data.get("k_a", [])
        num_experts = len(table_competency)

        if not table_competency:
            ws["A3"] = "Немає даних про компетентність"
            self.set_data_style(ws["A3"])
            return

        # Table headers
        ws["A3"] = "Експерти"
        self.set_subheader_style(ws["A3"])

        # Source arguments headers
        for i, arg in enumerate(name_arguments):
            col_letter = chr(66 + i)  # B, C, D, etc.
            ws[f"{col_letter}3"] = arg
            self.set_subheader_style(ws[f"{col_letter}3"])

        # K_k and K_a headers
        k_k_col = chr(66 + len(name_arguments))
        k_a_col = chr(67 + len(name_arguments))
        ws[f"{k_k_col}3"] = "K_k"
        ws[f"{k_a_col}3"] = "K_a"
        self.set_subheader_style(ws[f"{k_k_col}3"])
        self.set_subheader_style(ws[f"{k_a_col}3"])

        # Data rows
        for expert in range(num_experts):
            row = expert + 4
            ws[f"A{row}"] = f"Експерт {expert + 1}"
            self.set_data_style(ws[f"A{row}"])

            # Competency data
            for i, arg in enumerate(name_arguments):
                col_letter = chr(66 + i)
                value = (
                    table_competency[expert][i]
                    if expert < len(table_competency)
                    and i < len(table_competency[expert])
                    else 0
                )
                self.set_number_style(ws[f"{col_letter}{row}"], value)

            # K_k and K_a values
            if expert < len(k_k):
                self.set_number_style(ws[f"{k_k_col}{row}"], k_k[expert])
            if expert < len(k_a):
                self.set_number_style(ws[f"{k_a_col}{row}"], k_a[expert])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_experts_data_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Дані експертів")

        # Header
        ws["A1"] = "Результати обчислень"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:E1")

        experts_data_table = analysis_data.get("experts_data_table", [])
        name_research = analysis_data.get("name_research", [])
        m_i = analysis_data.get("m_i", [])
        r_i = analysis_data.get("r_i", [])
        l_value = analysis_data.get("l_value", [])
        l_value_sum = analysis_data.get("l_value_sum", 0)
        num_experts = len(experts_data_table)

        if not experts_data_table:
            ws["A3"] = "Немає даних експертів"
            self.set_data_style(ws["A3"])
            return

        # Table headers
        ws["A3"] = "Експерти"
        self.set_subheader_style(ws["A3"])

        # Research criteria headers
        for i, research in enumerate(name_research):
            col_letter = chr(66 + i)  # B, C, D, etc.
            ws[f"{col_letter}3"] = research
            self.set_subheader_style(ws[f"{col_letter}3"])

        # Data rows
        for expert in range(num_experts):
            row = expert + 4
            ws[f"A{row}"] = f"Експерт {expert + 1}"
            self.set_data_style(ws[f"A{row}"])

            # Expert data
            for i, research in enumerate(name_research):
                col_letter = chr(66 + i)
                value = (
                    experts_data_table[expert][i]
                    if expert < len(experts_data_table)
                    and i < len(experts_data_table[expert])
                    else 0
                )
                self.set_number_style(ws[f"{col_letter}{row}"], value)

        # Calculation rows
        calc_row = num_experts + 4

        # M_i row
        ws[f"A{calc_row}"] = "M_i"
        self.set_sum_style(ws[f"A{calc_row}"])
        for i, research in enumerate(name_research):
            col_letter = chr(66 + i)
            value = m_i[i] if i < len(m_i) else 0
            self.set_number_style(ws[f"{col_letter}{calc_row}"], value)

        # R_i row
        calc_row += 1
        ws[f"A{calc_row}"] = "R_i"
        self.set_sum_style(ws[f"A{calc_row}"])
        for i, research in enumerate(name_research):
            col_letter = chr(66 + i)
            value = r_i[i] if i < len(r_i) else 0
            self.set_number_style(ws[f"{col_letter}{calc_row}"], value)

        # Lambda row
        calc_row += 1
        ws[f"A{calc_row}"] = "λ"
        self.set_sum_style(ws[f"A{calc_row}"])
        for i, research in enumerate(name_research):
            col_letter = chr(66 + i)
            value = l_value[i] if i < len(l_value) else 0
            self.set_number_style(ws[f"{col_letter}{calc_row}"], value)

        # Sum column
        sum_col = chr(66 + len(name_research))
        ws[f"{sum_col}{calc_row}"] = l_value_sum
        self.set_sum_style(ws[f"{sum_col}{calc_row}"])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_ranking_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Ранжування")

        # Header
        ws["A1"] = "Ранжування результатів"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:B1")

        # Table headers
        ws["A3"] = "Критерій"
        ws["B3"] = "Значення λ"
        self.set_subheader_style(ws["A3"])
        self.set_subheader_style(ws["B3"])

        name_research = analysis_data.get("name_research", [])
        l_value = analysis_data.get("l_value", [])

        if not name_research or not l_value:
            ws["A5"] = "Немає даних для ранжування"
            self.set_data_style(ws["A5"])
            return

        # Create ranking data
        ranking_data = []
        for i, research in enumerate(name_research):
            value = l_value[i] if i < len(l_value) else 0
            ranking_data.append((research, value))

        # Sort by value in descending order
        ranking_data.sort(key=lambda x: x[1], reverse=True)

        # Fill ranking table
        for i, (research, value) in enumerate(ranking_data):
            row = i + 5
            ws[f"A{row}"] = research
            ws[f"B{row}"] = round(value, 3)

            self.set_data_style(ws[f"A{row}"])
            self.set_number_style(ws[f"B{row}"], value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_chart_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Графіки")

        # Header
        ws["A1"] = "Візуалізація результатів"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:D1")

        name_research = analysis_data.get("name_research", [])
        l_value = analysis_data.get("l_value", [])

        if not name_research or not l_value:
            ws["A3"] = "Немає даних для графіка"
            self.set_data_style(ws["A3"])
            return

        # Prepare data for chart
        ws["A3"] = "Критерій"
        ws["B3"] = "Значення λ"
        self.set_subheader_style(ws["A3"])
        self.set_subheader_style(ws["B3"])

        for i, research in enumerate(name_research):
            row = i + 4
            ws[f"A{row}"] = research
            ws[f"B{row}"] = l_value[i] if i < len(l_value) else 0
            self.set_data_style(ws[f"A{row}"])
            self.set_number_style(ws[f"B{row}"], l_value[i] if i < len(l_value) else 0)

        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Результати експертних оцінок"
        chart.y_axis.title = "Значення λ"
        chart.x_axis.title = "Критерії"

        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(name_research))
        categories = Reference(ws, min_col=1, min_row=4, max_row=3 + len(name_research))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "E3")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_conclusion_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Висновки")

        # Header
        ws["A1"] = "Висновки"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:D1")

        rank_str = analysis_data.get("rank_str", "Аналіз завершено")

        ws["A3"] = "Результат аналізу:"
        self.set_subheader_style(ws["A3"])

        ws["A4"] = rank_str
        self.set_data_style(ws["A4"])
        ws.merge_cells("A4:D4")
        self.set_row_height_for_text(ws, 4, rank_str)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def generate_experts_analysis_excel(self, analysis_data):
        """Generate complete Excel file for experts analysis"""

        # Create sheets in order
        self.create_general_info_sheet(analysis_data)
        self.create_competency_sheet(analysis_data)
        self.create_experts_data_sheet(analysis_data)
        self.create_ranking_sheet(analysis_data)
        self.create_chart_sheet(analysis_data)
        self.create_conclusion_sheet(analysis_data)

        return self.workbook

    def save_to_bytes(self):
        """Save workbook to bytes"""
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
