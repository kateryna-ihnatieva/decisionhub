#!/usr/bin/env python3
"""
Binary Relations Excel Export Module
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image


class BinaryExcelExporter:
    """Excel exporter for Binary Relations method"""

    def __init__(self):
        self.workbook = None
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"  # Dark blue
        )
        self.subheader_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"  # Medium blue
        )
        self.data_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"  # Light gray
        )
        self.sum_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"  # Light blue
        )
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_header_style(self, cell):
        """Set header cell style"""
        cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        cell.fill = self.header_fill
        cell.border = self.border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_subheader_style(self, cell):
        """Set subheader cell style"""
        cell.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        cell.fill = self.subheader_fill
        cell.border = self.border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_data_style(self, cell):
        """Set data cell style"""
        cell.font = Font(name="Arial", size=11, bold=False, color="000000")
        cell.fill = self.data_fill
        cell.border = self.border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_sum_style(self, cell):
        """Set sum cell style"""
        cell.font = Font(name="Arial", size=11, bold=True, color="000000")
        cell.fill = self.sum_fill
        cell.border = self.border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_number_style(self, cell, value):
        """Set number cell style with proper formatting"""
        cell.font = Font(name="Arial", size=11, bold=False, color="000000")
        cell.fill = self.data_fill
        cell.border = self.border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Handle string fractions
        if isinstance(value, str) and "/" in value:
            try:
                parts = value.split("/")
                if len(parts) == 2:
                    numerator = float(parts[0])
                    denominator = float(parts[1])
                    if denominator != 0:
                        cell.value = int(numerator / denominator)
                        cell.number_format = "0"
                    else:
                        cell.value = value
                else:
                    cell.value = value
            except (ValueError, ZeroDivisionError):
                cell.value = value
        else:
            cell.value = value
            if isinstance(value, (int, float)):
                cell.value = int(value)
                cell.number_format = "0"

    def auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max(max_length + 2, 15), 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def create_general_info_sheet(self, analysis_data):
        """Create general information sheet"""
        ws = self.workbook.create_sheet("Загальна інформація", 0)

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "Звіт аналізу бінарних відношень"
        self.set_header_style(title_cell)

        # Method info
        ws["A3"] = "Метод аналізу:"
        ws["B3"] = "Бінарні відношення"
        self.set_data_style(ws["A3"])
        self.set_data_style(ws["B3"])

        ws["A4"] = "ID аналізу:"
        ws["B4"] = analysis_data.get("method_id", "N/A")
        self.set_data_style(ws["A4"])
        self.set_data_style(ws["B4"])

        ws["A5"] = "Дата створення:"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.set_data_style(ws["A5"])
        self.set_data_style(ws["B5"])

        ws["A6"] = "Опис завдання:"
        ws["B6"] = analysis_data.get("task_description", "Не вказано")
        self.set_data_style(ws["A6"])
        self.set_data_style(ws["B6"])

        self.auto_adjust_columns(ws)

    def create_matrix_sheet(self, analysis_data):
        """Create matrix comparison sheet"""
        ws = self.workbook.create_sheet("Матриця порівнянь")

        names = analysis_data["names"]
        matrix = analysis_data["matrix"]
        sorted_dict = analysis_data["sorted_dict"]

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "Матриця попарних порівнянь"
        self.set_header_style(title_cell)

        # Headers
        ws["A3"] = ""
        for i, name in enumerate(names):
            col_letter = get_column_letter(i + 2)
            ws[f"{col_letter}3"] = name
            self.set_subheader_style(ws[f"{col_letter}3"])

        ws[f"{get_column_letter(len(names) + 2)}3"] = "Сума"
        self.set_subheader_style(ws[f"{get_column_letter(len(names) + 2)}3"])

        # Matrix data
        for i, name in enumerate(names):
            row = i + 4
            ws[f"A{row}"] = name
            self.set_subheader_style(ws[f"A{row}"])

            for j in range(len(names)):
                col_letter = get_column_letter(j + 2)
                self.set_number_style(ws[f"{col_letter}{row}"], matrix[i][j])

            # Sum column
            sum_col = get_column_letter(len(names) + 2)
            ws[f"{sum_col}{row}"] = sorted_dict[name]
            self.set_sum_style(ws[f"{sum_col}{row}"])

        self.auto_adjust_columns(ws)

    def create_ranking_sheet(self, analysis_data):
        """Create ranking results sheet"""
        ws = self.workbook.create_sheet("Ранжування")

        sorted_dict = analysis_data["sorted_dict"]
        ranj_str = analysis_data["ranj_str"]

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "Результати ранжування"
        self.set_header_style(title_cell)

        # Ranking order
        ws["A3"] = "Порядок ранжування:"
        ws["B3"] = ranj_str
        self.set_data_style(ws["A3"])
        self.set_data_style(ws["B3"])

        # Ranking table
        ws["A5"] = "Об'єкт"
        ws["B5"] = "Сума"
        self.set_subheader_style(ws["A5"])
        self.set_subheader_style(ws["B5"])

        # Data - sorted by value in descending order
        sorted_items = sorted(sorted_dict.items(), key=lambda x: x[1], reverse=True)
        for i, (name, total) in enumerate(sorted_items):
            row = i + 6
            ws[f"A{row}"] = name
            ws[f"B{row}"] = int(total)
            self.set_data_style(ws[f"A{row}"])
            self.set_sum_style(ws[f"B{row}"])

        self.auto_adjust_columns(ws)

    def create_transitivity_sheet(self, analysis_data):
        """Create transitivity check sheet"""
        ws = self.workbook.create_sheet("Транзитивність")

        comb = analysis_data["comb"]
        cond_tranz = analysis_data["cond_tranz"]
        vidnosh = analysis_data["vidnosh"]
        prim = analysis_data["prim"]

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "Перевірка на транзитивність"
        self.set_header_style(title_cell)

        # Headers
        headers = [
            "№",
            "Комбінація 3-х об'єктів",
            "Умова транзитивності",
            "Відношення",
            "Примітка",
        ]
        for i, header in enumerate(headers):
            col_letter = get_column_letter(i + 1)
            ws[f"{col_letter}3"] = header
            self.set_subheader_style(ws[f"{col_letter}3"])

        # Data
        for i in range(len(comb)):
            row = i + 4
            ws[f"A{row}"] = i + 1
            ws[f"B{row}"] = comb[i]
            ws[f"C{row}"] = cond_tranz[i]
            ws[f"D{row}"] = vidnosh[i]

            # Status cell with special formatting
            status_cell = ws[f"E{row}"]
            status_cell.value = prim[i]
            status_cell.font = Font(size=11, bold=True, color="FFFFFF")
            status_cell.border = self.border
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

            if prim[i] == "+":
                status_cell.fill = PatternFill(
                    start_color="32CD32", end_color="32CD32", fill_type="solid"
                )
            elif prim[i] == "-":
                status_cell.fill = PatternFill(
                    start_color="DC143C", end_color="DC143C", fill_type="solid"
                )
            else:
                status_cell.fill = self.data_fill

            # Other cells
            self.set_data_style(ws[f"A{row}"])
            self.set_data_style(ws[f"B{row}"])
            self.set_data_style(ws[f"C{row}"])
            self.set_data_style(ws[f"D{row}"])

        self.auto_adjust_columns(ws)

    def create_formal_notation_sheet(self, analysis_data):
        """Create formal notation sheet"""
        ws = self.workbook.create_sheet("Формальні позначки")

        names = analysis_data["names"]

        # Title
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = "Формальні позначки об'єктів"
        self.set_header_style(title_cell)

        # Headers
        ws["A3"] = "Об'єкт"
        ws["B3"] = "Формальна позначка"
        self.set_subheader_style(ws["A3"])
        self.set_subheader_style(ws["B3"])

        # Data
        for i, name in enumerate(names):
            row = i + 4
            ws[f"A{row}"] = name
            ws[f"B{row}"] = f"a{i+1}"
            self.set_data_style(ws[f"A{row}"])
            self.set_data_style(ws[f"B{row}"])

        self.auto_adjust_columns(ws)

    def create_chart_sheet(self, analysis_data):
        """Create charts sheet"""
        ws = self.workbook.create_sheet("Графіки")

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "Візуалізація результатів"
        self.set_header_style(title_cell)

        # Create ranking chart
        self.create_ranking_chart(ws, analysis_data)

        self.auto_adjust_columns(ws)

    def create_ranking_chart(self, ws, analysis_data):
        """Create ranking bar chart using openpyxl"""
        sorted_dict = analysis_data["sorted_dict"]

        # Prepare data - sorted by value in descending order
        sorted_items = sorted(sorted_dict.items(), key=lambda x: x[1], reverse=True)
        names = [item[0] for item in sorted_items]
        values = [int(item[1]) for item in sorted_items]

        # Add data to worksheet
        ws["A3"] = "Об'єкт"
        ws["B3"] = "Сума балів"
        self.set_subheader_style(ws["A3"])
        self.set_subheader_style(ws["B3"])

        for i, (name, value) in enumerate(sorted_items):
            row = i + 4
            ws[f"A{row}"] = name
            ws[f"B{row}"] = int(value)
            self.set_data_style(ws[f"A{row}"])
            self.set_data_style(ws[f"B{row}"])

        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Ранжування об'єктів"
        chart.y_axis.title = "Сума балів"
        chart.x_axis.title = "Об'єкти"

        # Set data range
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(names))
        categories = Reference(ws, min_col=1, min_row=4, max_row=3 + len(names))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, "E3")

    def create_conclusion_sheet(self, analysis_data):
        """Create conclusion sheet"""
        ws = self.workbook.create_sheet("Висновки")

        visnovok = analysis_data.get("visnovok", "Висновок не вказано")

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "Висновки аналізу"
        self.set_header_style(title_cell)

        # Conclusion text
        ws["A3"] = "Висновок:"
        ws["A3"].font = Font(size=13, bold=True, color="000000")
        ws["A3"].alignment = Alignment(vertical="top")

        # Wrap conclusion text
        ws["A4"] = visnovok
        ws["A4"].font = Font(size=11, color="000000")
        ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A4:D20")

        self.auto_adjust_columns(ws)

    def generate_binary_analysis_excel(self, analysis_data):
        """Generate complete Excel file for binary analysis"""
        self.workbook = Workbook()

        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        # Create sheets in order
        self.create_general_info_sheet(analysis_data)
        self.create_matrix_sheet(analysis_data)
        self.create_ranking_sheet(analysis_data)
        self.create_transitivity_sheet(analysis_data)
        self.create_formal_notation_sheet(analysis_data)
        self.create_chart_sheet(analysis_data)
        self.create_conclusion_sheet(analysis_data)

        return self.workbook

    def save_to_bytes(self):
        """Save workbook to bytes"""
        if not self.workbook:
            raise ValueError("No workbook to save")

        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
