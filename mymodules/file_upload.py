import os
import pandas as pd
import openpyxl
from werkzeug.utils import secure_filename
from flask import current_app


class FileUploadError(Exception):
    """Custom exception for file upload errors"""

    pass


def validate_file(file, max_size_mb=10):
    """
    Validate uploaded file for format and size

    Args:
        file: Flask file object
        max_size_mb: Maximum file size in MB

    Returns:
        tuple: (is_valid, error_message)
    """
    if not file or not file.filename:
        return False, "No file selected"

    # Check file extension
    allowed_extensions = {".xlsx", ".xls", ".csv"}
    file_ext = os.path.splitext(file.filename.lower())[1]

    if file_ext not in allowed_extensions:
        formats = ", ".join(allowed_extensions)
        return (
            False,
            f"Unsupported file format. Allowed formats: {formats}",
        )

    # Check file size
    file.seek(0, 2)  # Seek to end
    file_size = file.tell()
    file.seek(0)  # Reset to beginning

    max_size_bytes = max_size_mb * 1024 * 1024
    if file_size > max_size_bytes:
        return False, f"File too large. Maximum size: {max_size_mb}MB"

    return True, None


def parse_excel_file(file_path, expected_size):
    """
    Parse Excel file and extract names and matrix data

    Args:
        file_path: Path to Excel file
        expected_size: Expected matrix size (rows/columns)

    Returns:
        dict: {'names': list, 'matrix': list, 'error': str or None}
    """
    try:
        # Try to read with openpyxl first (better for .xlsx)
        if file_path.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active

            # Convert to list of lists
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append([cell if cell is not None else "" for cell in row])
        else:
            # Use pandas for .xls files
            df = pd.read_excel(file_path, header=None)
            data = df.fillna("").values.tolist()

        return extract_names_and_matrix(data, expected_size)

    except Exception as e:
        return {
            "names": [],
            "matrix": [],
            "error": f"Error reading Excel file: {str(e)}",
        }


def parse_csv_file(file_path, expected_size):
    """
    Parse CSV file and extract names and matrix data

    Args:
        file_path: Path to CSV file
        expected_size: Expected matrix size (rows/columns)

    Returns:
        dict: {'names': list, 'matrix': list, 'error': str or None}
    """
    try:
        # Try different separators
        separators = [",", ";", "\t", "|"]
        data = None

        for sep in separators:
            try:
                df = pd.read_csv(file_path, header=None, sep=sep)
                if not df.empty:
                    data = df.fillna("").values.tolist()
                    break
            except Exception:
                continue

        if data is None:
            return {
                "names": [],
                "matrix": [],
                "error": "Could not parse CSV file with any known separator",
            }

        return extract_names_and_matrix(data, expected_size)

    except Exception as e:
        return {"names": [], "matrix": [], "error": f"Error reading CSV file: {str(e)}"}


def extract_names_and_matrix(data, expected_size):
    """
    Extract names and matrix from parsed data for hierarchy analysis

    Args:
        data: List of lists containing the file data
        expected_size: Expected matrix size (rows/columns)

    Returns:
        dict: {'names': list, 'matrix': list, 'error': str or None}
    """
    try:
        if not data or len(data) == 0:
            return {"names": [], "matrix": [], "error": "File is empty"}

        # Clean data - remove completely empty rows
        data = [row for row in data if any(str(cell).strip() for cell in row)]

        if len(data) == 0:
            return {"names": [], "matrix": [], "error": "No data found in file"}

        # Determine matrix structure
        rows = len(data)
        cols = len(data[0]) if data else 0

        # Check if we have enough data for the expected size
        min_size = expected_size + 1
        if rows < min_size or cols < min_size:
            return {
                "names": [],
                "matrix": [],
                "error": f"Insufficient data. Expected at least {min_size-1}x{min_size-1}, got {rows-1}x{cols-1}",
            }

        # For hierarchy analysis, we expect names in both first row and first column
        # First cell (0,0) is usually empty or contains a label
        # Names are in first row (excluding first cell) and first column

        # Extract names from first row (excluding first cell)
        first_row_names = []
        for j in range(1, min(expected_size + 1, len(data[0]))):
            cell_value = str(data[0][j]).strip()
            if cell_value and cell_value != "":
                first_row_names.append(cell_value)

        # Extract names from first column (excluding first cell)
        first_col_names = []
        for i in range(1, min(expected_size + 1, len(data))):
            cell_value = str(data[i][0]).strip()
            if cell_value and cell_value != "":
                first_col_names.append(cell_value)

        # Determine which names to use (prefer first row, fallback to first column)
        names = []
        matrix_start_row = 1
        matrix_start_col = 1

        if len(first_row_names) >= expected_size:
            names = first_row_names[:expected_size]
        elif len(first_col_names) >= expected_size:
            names = first_col_names[:expected_size]
        else:
            # Try to combine both if neither has enough
            combined_names = list(set(first_row_names + first_col_names))
            if len(combined_names) >= expected_size:
                names = combined_names[:expected_size]
            else:
                # Use default names if nothing works
                names = [f"Item {i+1}" for i in range(expected_size)]
                matrix_start_row = 0
                matrix_start_col = 0

        # Extract matrix data
        matrix = []
        for i in range(matrix_start_row, matrix_start_row + expected_size):
            row = []
            for j in range(matrix_start_col, matrix_start_col + expected_size):
                if i < len(data) and j < len(data[i]):
                    cell_value = data[i][j]
                else:
                    cell_value = ""

                # Convert to string and clean
                cell_str = str(cell_value).strip()

                # Try to convert to float for numeric values
                # Handle fractions like "1/2", "1/5" etc.
                try:
                    if "/" in cell_str:
                        # Handle fractions
                        numerator, denominator = cell_str.split("/")
                        row.append(float(numerator) / float(denominator))
                    elif (
                        cell_str
                        and cell_str.replace(".", "")
                        .replace("-", "")
                        .replace("+", "")
                        .replace("e", "")
                        .replace("E", "")
                        .isdigit()
                    ):
                        row.append(float(cell_str))
                    else:
                        # Try to evaluate as expression (for cases like "1/2" written as "1/2")
                        try:
                            row.append(float(eval(cell_str)))
                        except Exception:
                            row.append(cell_str)
                except Exception:
                    row.append(cell_str)

            matrix.append(row)

        # Validate names
        if not names or len(names) != expected_size:
            return {
                "names": [],
                "matrix": [],
                "error": f"Could not extract {expected_size} valid names",
            }

        # Check for duplicate names
        if len(set(names)) != len(names):
            return {
                "names": [],
                "matrix": [],
                "error": "Duplicate names found. All names must be unique",
            }

        # Validate matrix contains numeric data
        numeric_count = 0
        for row in matrix:
            for cell in row:
                if isinstance(cell, (int, float)) and not pd.isna(cell):
                    numeric_count += 1

        if numeric_count == 0:
            return {
                "names": [],
                "matrix": [],
                "error": "No numeric data found in matrix",
            }

        return {"names": names, "matrix": matrix, "error": None}

    except Exception as e:
        return {"names": [], "matrix": [], "error": f"Error processing data: {str(e)}"}


def save_uploaded_file(file, upload_folder):
    """
    Save uploaded file to specified folder

    Args:
        file: Flask file object
        upload_folder: Path to upload folder

    Returns:
        str: Path to saved file
    """
    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)

    filename = secure_filename(file.filename)
    file_path = os.path.join(upload_folder, filename)
    file.save(file_path)

    return file_path


def cleanup_file(file_path):
    """
    Delete uploaded file after processing

    Args:
        file_path: Path to file to delete
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        current_app.logger.warning(f"Could not delete file {file_path}: {str(e)}")


def process_uploaded_file(file, expected_size, upload_folder="uploads"):
    """
    Main function to process uploaded file

    Args:
        file: Flask file object
        expected_size: Expected matrix size
        upload_folder: Folder to save uploaded files

    Returns:
        dict: {'success': bool, 'names': list, 'matrix': list, 'error': str}
    """
    try:
        # Validate file
        is_valid, error = validate_file(file)
        if not is_valid:
            return {"success": False, "names": [], "matrix": [], "error": error}

        # Save file
        file_path = save_uploaded_file(file, upload_folder)

        try:
            # Parse file based on extension
            file_ext = os.path.splitext(file.filename.lower())[1]

            if file_ext in [".xlsx", ".xls"]:
                result = parse_excel_file(file_path, expected_size)
            elif file_ext == ".csv":
                result = parse_csv_file(file_path, expected_size)
            else:
                return {
                    "success": False,
                    "names": [],
                    "matrix": [],
                    "error": "Unsupported file format",
                }

            # Clean up file
            cleanup_file(file_path)

            if result["error"]:
                return {
                    "success": False,
                    "names": [],
                    "matrix": [],
                    "error": result["error"],
                }

            return {
                "success": True,
                "names": result["names"],
                "matrix": result["matrix"],
                "error": None,
            }

        except Exception as e:
            # Clean up file on error
            cleanup_file(file_path)
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": f"Error processing file: {str(e)}",
            }

    except Exception as e:
        return {
            "success": False,
            "names": [],
            "matrix": [],
            "error": f"Upload error: {str(e)}",
        }


def process_hierarchy_file(
    file, num_criteria, num_alternatives, upload_folder="uploads"
):
    """
    Process uploaded file specifically for hierarchy analysis
    Handles files with both criteria and alternatives matrices

    Args:
        file: Flask file object
        num_criteria: Number of criteria
        num_alternatives: Number of alternatives
        upload_folder: Folder to save uploaded files

    Returns:
        dict: {'success': bool, 'criteria_names': list, 'alternatives_names': list,
               'criteria_matrix': list, 'alternatives_matrices': list, 'error': str}
    """
    try:
        # Validate file
        is_valid, error = validate_file(file)
        if not is_valid:
            return {
                "success": False,
                "criteria_names": [],
                "alternatives_names": [],
                "criteria_matrix": [],
                "alternatives_matrices": [],
                "error": error,
            }

        # Save file
        file_path = save_uploaded_file(file, upload_folder)

        try:
            # Parse file based on extension
            file_ext = os.path.splitext(file.filename.lower())[1]

            if file_ext in [".xlsx", ".xls"]:
                data = parse_excel_data(file_path)
            elif file_ext == ".csv":
                data = parse_csv_data_with_empty_rows(file_path)
            else:
                return {
                    "success": False,
                    "criteria_names": [],
                    "alternatives_names": [],
                    "criteria_matrix": [],
                    "alternatives_matrices": [],
                    "error": "Unsupported file format",
                }

            # Process hierarchy data
            result = extract_hierarchy_data(data, num_criteria, num_alternatives)

            # Clean up file
            cleanup_file(file_path)

            return result

        except Exception as e:
            # Clean up file on error
            cleanup_file(file_path)
            return {
                "success": False,
                "criteria_names": [],
                "alternatives_names": [],
                "criteria_matrix": [],
                "alternatives_matrices": [],
                "error": f"Error processing file: {str(e)}",
            }

    except Exception as e:
        return {
            "success": False,
            "criteria_names": [],
            "alternatives_names": [],
            "criteria_matrix": [],
            "alternatives_matrices": [],
            "error": f"Upload error: {str(e)}",
        }


def parse_excel_data(file_path):
    """Parse Excel file and return raw data"""
    if file_path.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append([cell if cell is not None else "" for cell in row])
    else:
        df = pd.read_excel(file_path, header=None)
        data = df.fillna("").values.tolist()
    return data


def parse_csv_data(file_path):
    """Parse CSV file and return raw data"""
    separators = [",", ";", "\t", "|"]
    data = None

    for sep in separators:
        try:
            df = pd.read_csv(file_path, header=None, sep=sep)
            if not df.empty:
                data = df.fillna("").values.tolist()
                break
        except Exception:
            continue

    if data is None:
        raise Exception("Could not parse CSV file with any known separator")

    return data


def parse_csv_data_with_empty_rows(file_path):
    """Parse CSV file and return raw data preserving empty rows"""
    separators = [",", ";", "\t", "|"]
    data = None

    for sep in separators:
        try:
            df = pd.read_csv(
                file_path,
                header=None,
                sep=sep,
                keep_default_na=False,
                dtype=str,
                skip_blank_lines=False,
            )
            if not df.empty:
                data = df.values.tolist()
                break
        except Exception:
            continue

    if data is None:
        raise Exception("Could not parse CSV file with any known separator")

    return data


def extract_hierarchy_data(data, num_criteria, num_alternatives):
    """
    Extract criteria and alternatives data from hierarchy file

    For hierarchy analysis:
    - First matrix: criteria names and matrix
    - Second matrix: alternatives names and matrix
    - Subsequent matrices: alternatives comparison matrices for each criterion

    Args:
        data: Raw file data
        num_criteria: Number of criteria
        num_alternatives: Number of alternatives

    Returns:
        dict with criteria and alternatives data
    """
    try:
        if not data or len(data) == 0:
            return {
                "success": False,
                "criteria_names": [],
                "alternatives_names": [],
                "criteria_matrix": [],
                "alternatives_matrices": [],
                "error": "File is empty",
            }

        # Find matrix boundaries by looking for empty rows
        matrix_boundaries = []
        current_start = 0

        for i, row in enumerate(data):
            # Check if row is completely empty
            if not any(str(cell).strip() for cell in row):
                if i > current_start:
                    matrix_boundaries.append((current_start, i))
                current_start = i + 1

        # Add the last matrix if it doesn't end with empty row
        if current_start < len(data):
            matrix_boundaries.append((current_start, len(data)))

        # If no empty rows found, treat the entire data as one matrix
        if not matrix_boundaries:
            matrix_boundaries.append((0, len(data)))

        if len(matrix_boundaries) < 2:
            return {
                "success": False,
                "criteria_names": [],
                "alternatives_names": [],
                "criteria_matrix": [],
                "alternatives_matrices": [],
                "error": "File must contain at least 2 matrices (criteria and alternatives)",
            }

        # Clean data for each matrix separately
        cleaned_matrices = []
        for start, end in matrix_boundaries:
            matrix_data = data[start:end]
            # Clean data - remove completely empty rows
            cleaned_matrix = [
                row for row in matrix_data if any(str(cell).strip() for cell in row)
            ]
            if cleaned_matrix:  # Only add non-empty matrices
                # Split each row by comma if it's a single string
                processed_matrix = []
                for row in cleaned_matrix:
                    if len(row) == 1 and isinstance(row[0], str) and "," in row[0]:
                        # Split the string by comma
                        processed_row = [cell.strip() for cell in row[0].split(",")]
                        processed_matrix.append(processed_row)
                    else:
                        processed_matrix.append(row)
                cleaned_matrices.append(processed_matrix)

        if len(cleaned_matrices) < 2:
            return {
                "success": False,
                "criteria_names": [],
                "alternatives_names": [],
                "criteria_matrix": [],
                "alternatives_matrices": [],
                "error": "File must contain at least 2 matrices (criteria and alternatives)",
            }

        # Extract first matrix (criteria)
        criteria_data = cleaned_matrices[0]
        criteria_result = extract_names_and_matrix(criteria_data, num_criteria)

        if criteria_result["error"]:
            return {
                "success": False,
                "criteria_names": [],
                "alternatives_names": [],
                "criteria_matrix": [],
                "alternatives_matrices": [],
                "error": f"Error extracting criteria: {criteria_result['error']}",
            }

        # Extract second matrix (alternatives)
        alternatives_data = cleaned_matrices[1]
        alternatives_result = extract_names_and_matrix(
            alternatives_data, num_alternatives
        )

        if alternatives_result["error"]:
            return {
                "success": False,
                "criteria_names": [],
                "alternatives_names": [],
                "criteria_matrix": [],
                "alternatives_matrices": [],
                "error": f"Error extracting alternatives: {alternatives_result['error']}",
            }

        # For hierarchy analysis, we need multiple alternatives matrices
        # Extract all matrices after the first one (criteria matrix)
        alternatives_matrices = []

        # Process all matrices after the first one
        for i in range(1, len(cleaned_matrices)):
            alt_matrix_data = cleaned_matrices[i]
            alt_result = extract_names_and_matrix(alt_matrix_data, num_alternatives)

            if alt_result["error"]:
                return {
                    "success": False,
                    "criteria_names": [],
                    "alternatives_names": [],
                    "criteria_matrix": [],
                    "alternatives_matrices": [],
                    "error": f"Error extracting alternatives matrix {i}: {alt_result['error']}",
                }

            alternatives_matrices.append(alt_result["matrix"])

        # Validate that we have the right number of matrices
        num_criteria = len(criteria_result["names"])
        if len(alternatives_matrices) != num_criteria:
            return {
                "success": False,
                "criteria_names": [],
                "alternatives_names": [],
                "criteria_matrix": [],
                "alternatives_matrices": [],
                "error": f"Expected {num_criteria} alternatives matrices, but found {len(alternatives_matrices)}",
            }

        return {
            "success": True,
            "criteria_names": criteria_result["names"],
            "alternatives_names": alternatives_result["names"],
            "criteria_matrix": criteria_result["matrix"],
            "alternatives_matrices": alternatives_matrices,
            "error": None,
        }

    except Exception as e:
        return {
            "success": False,
            "criteria_names": [],
            "alternatives_names": [],
            "criteria_matrix": [],
            "alternatives_matrices": [],
            "error": f"Error processing hierarchy data: {str(e)}",
        }


def process_binary_file(file, num_objects, upload_folder="uploads"):
    """
    Process uploaded file specifically for binary relations analysis
    Handles files with binary relations matrix

    Args:
        file: Flask file object
        num_objects: Number of objects for binary relations
        upload_folder: Folder to save uploaded files

    Returns:
        dict: {'success': bool, 'names': list, 'matrix': list, 'error': str}
    """
    try:
        # Validate file
        is_valid, error = validate_file(file)
        if not is_valid:
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": error,
            }

        # Save file
        file_path = save_uploaded_file(file, upload_folder)

        try:
            # Parse file based on extension
            file_ext = os.path.splitext(file.filename.lower())[1]

            if file_ext in [".xlsx", ".xls"]:
                data = parse_excel_data(file_path)
            elif file_ext == ".csv":
                data = parse_csv_data_with_empty_rows(file_path)
            else:
                return {
                    "success": False,
                    "names": [],
                    "matrix": [],
                    "error": "Unsupported file format",
                }

            # Process binary relations data
            result = extract_binary_data(data, num_objects)

            # Clean up file
            cleanup_file(file_path)

            return result

        except Exception as e:
            # Clean up file on error
            cleanup_file(file_path)
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": f"Error processing file: {str(e)}",
            }

    except Exception as e:
        return {
            "success": False,
            "names": [],
            "matrix": [],
            "error": f"Upload error: {str(e)}",
        }


def extract_binary_data(data, num_objects):
    """
    Extract binary relations data from file

    For binary relations analysis:
    - Matrix with object names in first row and first column
    - Values should be -1, 0, or 1
    - Matrix should be square (num_objects x num_objects)

    Args:
        data: Raw file data
        num_objects: Number of objects

    Returns:
        dict with names and matrix data
    """
    try:
        if not data or len(data) == 0:
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": "File is empty",
            }

        # Clean data - remove completely empty rows
        data = [row for row in data if any(str(cell).strip() for cell in row)]

        if len(data) == 0:
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": "No data found in file",
            }

        # Check if we have enough data for the expected size
        min_size = num_objects + 1  # +1 for header row/column
        if len(data) < min_size or len(data[0]) < min_size:
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": f"Insufficient data. Expected at least {min_size-1}x{min_size-1}, got {len(data)-1}x{len(data[0])-1}",
            }
        elif len(data) > min_size or len(data[0]) > min_size:
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": f"Excessive data. Expected at most {min_size-1}x{min_size-1}, got {len(data)-1}x{len(data[0])-1}",
            }

        # Extract names from first row (excluding first cell)
        names = []
        for j in range(1, min(num_objects + 1, len(data[0]))):
            cell_value = str(data[0][j]).strip()
            if cell_value and cell_value != "":
                names.append(cell_value)

        # If we don't have enough names from first row, try first column
        if len(names) < num_objects:
            for i in range(1, min(num_objects + 1, len(data))):
                cell_value = str(data[i][0]).strip()
                if cell_value and cell_value != "" and cell_value not in names:
                    names.append(cell_value)

        # Validate we have the correct number of names
        if len(names) != num_objects:
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": f"Expected {num_objects} object names, found {len(names)}",
            }

        # Take only the required number of names
        names = names[:num_objects]

        # Extract matrix data
        matrix = []
        for i in range(1, min(num_objects + 1, len(data))):
            matrix_row = []
            for j in range(1, min(num_objects + 1, len(data[i]))):
                cell_value = str(data[i][j]).strip()
                try:
                    # Convert to integer and validate it's -1, 0, or 1
                    value = int(float(cell_value))
                    if value not in [-1, 0, 1]:
                        return {
                            "success": False,
                            "names": [],
                            "matrix": [],
                            "error": f"Invalid value '{value}' at position ({i},{j}). Only -1, 0, and 1 are allowed.",
                        }
                    matrix_row.append(value)
                except (ValueError, TypeError):
                    return {
                        "success": False,
                        "names": [],
                        "matrix": [],
                        "error": f"Invalid value '{cell_value}' at position ({i},{j}). Must be a number.",
                    }

            # Pad row if necessary
            while len(matrix_row) < num_objects:
                matrix_row.append(0)
            matrix.append(matrix_row)

        # Pad matrix if necessary
        while len(matrix) < num_objects:
            matrix.append([0] * num_objects)

        # Validate matrix dimensions
        if len(matrix) != num_objects or len(matrix[0]) != num_objects:
            return {
                "success": False,
                "names": [],
                "matrix": [],
                "error": f"Matrix should be {num_objects}x{num_objects}, got {len(matrix)}x{len(matrix[0])}",
            }

        return {
            "success": True,
            "names": names,
            "matrix": matrix,
            "error": None,
        }

    except Exception as e:
        return {
            "success": False,
            "names": [],
            "matrix": [],
            "error": f"Error processing binary relations data: {str(e)}",
        }


def process_laplasa_file(
    file, num_alternatives, num_conditions, upload_folder="uploads"
):
    """
    Process uploaded file for Laplasa method

    Args:
        file: Uploaded file object
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions
        upload_folder: Folder to save uploaded files

    Returns:
        dict: {'success': bool, 'alternatives_names': list, 'conditions_names': list,
               'cost_matrix': list, 'error': str}
    """
    try:
        # Save uploaded file
        file_path = save_uploaded_file(file, upload_folder)

        # Parse file based on extension
        if file_path.endswith((".xlsx", ".xls")):
            data = pd.read_excel(file_path, header=None).values.tolist()
        elif file_path.endswith(".csv"):
            data = pd.read_csv(file_path, header=None).values.tolist()
        else:
            return {
                "success": False,
                "alternatives_names": [],
                "conditions_names": [],
                "cost_matrix": [],
                "error": "Unsupported file format",
            }

        # Extract data
        result = extract_laplasa_data(data, num_alternatives, num_conditions)

        # Clean up temporary file
        os.remove(file_path)

        return result

    except Exception as e:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Error processing file: {str(e)}",
        }


def extract_laplasa_data(data, num_alternatives, num_conditions):
    """
    Extract Laplasa data from parsed file data

    Args:
        data: Parsed file data (list of rows)
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions

    Returns:
        dict with alternatives names, conditions names, and cost matrix
    """
    if not data or len(data) < 2:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "No data found in file",
        }

    # Find table boundaries by looking for empty or mostly empty rows
    table_boundaries = []
    current_start = 0

    for i, row in enumerate(data):
        # Check if row is empty or mostly empty (less than 2 non-empty cells)
        non_empty_cells = sum(
            1
            for cell in row
            if str(cell).strip() and str(cell).strip().lower() != "nan"
        )
        if non_empty_cells < 2:  # Row is empty or has only 1 non-empty cell
            if i > current_start:
                table_boundaries.append((current_start, i))
            current_start = i + 1

    # Add the last table if it doesn't end with empty row
    if current_start < len(data):
        table_boundaries.append((current_start, len(data)))

    # If no empty rows found, treat the entire data as one table
    if not table_boundaries:
        table_boundaries = [(0, len(data))]

    if len(table_boundaries) < 1:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "File must contain at least 1 table",
        }

    # Process the table
    table_data = data[table_boundaries[0][0] : table_boundaries[0][1]]
    result = extract_laplasa_matrix(table_data, num_alternatives, num_conditions)

    if not result["success"]:
        return result

    return {
        "success": True,
        "alternatives_names": result["alternatives_names"],
        "conditions_names": result["conditions_names"],
        "cost_matrix": result["cost_matrix"],
        "error": None,
    }


def extract_laplasa_matrix(data, num_alternatives, num_conditions):
    """
    Extract Laplasa matrix from table data

    Args:
        data: Table data (list of rows)
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions

    Returns:
        dict with alternatives names, conditions names, and cost matrix
    """
    if len(data) < 2:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "Insufficient data for Laplasa matrix",
        }

    # First row should contain conditions names
    conditions_row = data[0]
    conditions_names = []
    for j in range(1, len(conditions_row)):  # Skip first column (alternatives)
        cell_value = str(conditions_row[j]).strip()
        if cell_value and cell_value.lower() != "nan":
            conditions_names.append(cell_value)
        else:
            conditions_names.append(f"Condition {j}")

    # Validate conditions count
    if len(conditions_names) != num_conditions:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Invalid number of conditions. Expected {num_conditions}, found {len(conditions_names)}",
        }

    # Extract alternatives names and cost matrix
    alternatives_names = []
    cost_matrix = []

    for i in range(1, len(data)):  # Skip header row
        row = data[i]
        if not row or len(row) == 0:
            continue

        # First column should contain alternative name
        alt_name = str(row[0]).strip()
        if alt_name and alt_name.lower() != "nan":
            alternatives_names.append(alt_name)
        else:
            alternatives_names.append(f"Alternative {i}")

        # Extract cost values
        cost_row = []
        for j in range(1, len(row)):  # Skip first column
            cell_value = str(row[j]).strip()
            if cell_value and cell_value.lower() != "nan":
                try:
                    # Handle checkmarks and other symbols - extract only numbers
                    import re

                    numbers = re.findall(r"-?\d+\.?\d*", cell_value)
                    if numbers:
                        value = float(numbers[0])
                        cost_row.append(value)
                    else:
                        return {
                            "success": False,
                            "alternatives_names": [],
                            "conditions_names": [],
                            "cost_matrix": [],
                            "error": f"Invalid cost value '{cell_value}' at position ({i},{j})",
                        }
                except (ValueError, TypeError):
                    return {
                        "success": False,
                        "alternatives_names": [],
                        "conditions_names": [],
                        "cost_matrix": [],
                        "error": f"Invalid cost value '{cell_value}' at position ({i},{j})",
                    }
            else:
                cost_row.append(0.0)

        # Pad row if necessary
        while len(cost_row) < num_conditions:
            cost_row.append(0.0)

        cost_matrix.append(cost_row)

    # Validate alternatives count
    if len(alternatives_names) != num_alternatives:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Invalid number of alternatives. Expected {num_alternatives}, found {len(alternatives_names)}",
        }

    return {
        "success": True,
        "alternatives_names": alternatives_names,
        "conditions_names": conditions_names,
        "cost_matrix": cost_matrix,
        "error": None,
    }


def process_maximin_file(
    file, num_alternatives, num_conditions, upload_folder="uploads"
):
    """
    Process uploaded file for Maximin method

    Args:
        file: Uploaded file object
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions
        upload_folder: Folder to save uploaded files

    Returns:
        dict: {'success': bool, 'alternatives_names': list, 'conditions_names': list,
               'cost_matrix': list, 'error': str}
    """
    try:
        # Save uploaded file
        file_path = save_uploaded_file(file, upload_folder)

        # Parse file based on extension
        if file_path.endswith((".xlsx", ".xls")):
            data = pd.read_excel(file_path, header=None).values.tolist()
        elif file_path.endswith(".csv"):
            data = pd.read_csv(file_path, header=None).values.tolist()
        else:
            return {
                "success": False,
                "alternatives_names": [],
                "conditions_names": [],
                "cost_matrix": [],
                "error": "Unsupported file format",
            }

        # Extract data
        result = extract_maximin_data(data, num_alternatives, num_conditions)

        # Clean up temporary file
        os.remove(file_path)

        return result

    except Exception as e:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Error processing file: {str(e)}",
        }


def extract_maximin_data(data, num_alternatives, num_conditions):
    """
    Extract Maximin data from parsed file data

    Args:
        data: Parsed file data (list of rows)
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions

    Returns:
        dict with alternatives names, conditions names, and cost matrix
    """
    if not data or len(data) < 2:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "No data found in file",
        }

    # Find table boundaries by looking for empty or mostly empty rows
    table_boundaries = []
    current_start = 0

    for i, row in enumerate(data):
        # Check if row is empty or mostly empty (less than 2 non-empty cells)
        non_empty_cells = sum(
            1
            for cell in row
            if str(cell).strip() and str(cell).strip().lower() != "nan"
        )
        if non_empty_cells < 2:  # Row is empty or has only 1 non-empty cell
            if i > current_start:
                table_boundaries.append((current_start, i))
            current_start = i + 1

    # Add the last table if it doesn't end with empty row
    if current_start < len(data):
        table_boundaries.append((current_start, len(data)))

    # If no empty rows found, treat the entire data as one table
    if not table_boundaries:
        table_boundaries = [(0, len(data))]

    if len(table_boundaries) < 1:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "File must contain at least 1 table",
        }

    # Process the table
    table_data = data[table_boundaries[0][0] : table_boundaries[0][1]]
    result = extract_maximin_matrix(table_data, num_alternatives, num_conditions)

    if not result["success"]:
        return result

    return {
        "success": True,
        "alternatives_names": result["alternatives_names"],
        "conditions_names": result["conditions_names"],
        "cost_matrix": result["cost_matrix"],
        "error": None,
    }


def extract_maximin_matrix(data, num_alternatives, num_conditions):
    """
    Extract Maximin matrix from table data

    Args:
        data: Table data (list of rows)
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions

    Returns:
        dict with alternatives names, conditions names, and cost matrix
    """
    if len(data) < 2:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "Insufficient data for Maximin matrix",
        }

    # First row should contain conditions names
    conditions_row = data[0]
    conditions_names = []
    for j in range(1, len(conditions_row)):  # Skip first column (alternatives)
        cell_value = str(conditions_row[j]).strip()
        if cell_value and cell_value.lower() != "nan":
            conditions_names.append(cell_value)
        else:
            conditions_names.append(f"Condition {j}")

    # Validate conditions count
    if len(conditions_names) != num_conditions:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Invalid number of conditions. Expected {num_conditions}, found {len(conditions_names)}",
        }

    # Extract alternatives names and cost matrix
    alternatives_names = []
    cost_matrix = []

    for i in range(1, len(data)):  # Skip header row
        row = data[i]
        if not row or len(row) == 0:
            continue

        # First column should contain alternative name
        alt_name = str(row[0]).strip()
        if alt_name and alt_name.lower() != "nan":
            alternatives_names.append(alt_name)
        else:
            alternatives_names.append(f"Alternative {i}")

        # Extract cost values
        cost_row = []
        for j in range(1, len(row)):  # Skip first column
            cell_value = str(row[j]).strip()
            if cell_value and cell_value.lower() != "nan":
                try:
                    # Handle checkmarks and other symbols - extract only numbers
                    import re

                    numbers = re.findall(r"-?\d+\.?\d*", cell_value)
                    if numbers:
                        value = float(numbers[0])
                        cost_row.append(value)
                    else:
                        return {
                            "success": False,
                            "alternatives_names": [],
                            "conditions_names": [],
                            "cost_matrix": [],
                            "error": f"Invalid cost value '{cell_value}' at position ({i},{j})",
                        }
                except (ValueError, TypeError):
                    return {
                        "success": False,
                        "alternatives_names": [],
                        "conditions_names": [],
                        "cost_matrix": [],
                        "error": f"Invalid cost value '{cell_value}' at position ({i},{j})",
                    }
            else:
                cost_row.append(0.0)

        # Pad row if necessary
        while len(cost_row) < num_conditions:
            cost_row.append(0.0)

        cost_matrix.append(cost_row)

    # Validate alternatives count
    if len(alternatives_names) != num_alternatives:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Invalid number of alternatives. Expected {num_alternatives}, found {len(alternatives_names)}",
        }

    return {
        "success": True,
        "alternatives_names": alternatives_names,
        "conditions_names": conditions_names,
        "cost_matrix": cost_matrix,
        "error": None,
    }


def process_savage_file(
    file, num_alternatives, num_conditions, upload_folder="uploads"
):
    """
    Process uploaded file for Savage method

    Args:
        file: Uploaded file object
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions
        upload_folder: Folder to save uploaded files

    Returns:
        dict: {'success': bool, 'alternatives_names': list, 'conditions_names': list,
               'cost_matrix': list, 'error': str}
    """
    try:
        # Save uploaded file
        file_path = save_uploaded_file(file, upload_folder)

        # Parse file based on extension
        if file_path.endswith((".xlsx", ".xls")):
            data = pd.read_excel(file_path, header=None).values.tolist()
        elif file_path.endswith(".csv"):
            data = pd.read_csv(file_path, header=None).values.tolist()
        else:
            return {
                "success": False,
                "alternatives_names": [],
                "conditions_names": [],
                "cost_matrix": [],
                "error": "Unsupported file format",
            }

        # Extract data
        result = extract_savage_data(data, num_alternatives, num_conditions)

        # Clean up temporary file
        os.remove(file_path)

        return result

    except Exception as e:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Error processing file: {str(e)}",
        }


def extract_savage_data(data, num_alternatives, num_conditions):
    """
    Extract Savage data from parsed file data

    Args:
        data: Parsed file data (list of rows)
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions

    Returns:
        dict with alternatives names, conditions names, and cost matrix
    """
    if not data or len(data) < 2:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "No data found in file",
        }

    # Find table boundaries by looking for empty or mostly empty rows
    table_boundaries = []
    current_start = 0

    for i, row in enumerate(data):
        # Check if row is empty or mostly empty (less than 2 non-empty cells)
        non_empty_cells = sum(
            1
            for cell in row
            if str(cell).strip() and str(cell).strip().lower() != "nan"
        )
        if non_empty_cells < 2:  # Row is empty or has only 1 non-empty cell
            if i > current_start:
                table_boundaries.append((current_start, i))
            current_start = i + 1

    # Add the last table if it doesn't end with empty row
    if current_start < len(data):
        table_boundaries.append((current_start, len(data)))

    # If no empty rows found, treat the entire data as one table
    if not table_boundaries:
        table_boundaries = [(0, len(data))]

    if len(table_boundaries) < 1:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "File must contain at least 1 table",
        }

    # Process the table
    table_data = data[table_boundaries[0][0] : table_boundaries[0][1]]
    result = extract_savage_matrix(table_data, num_alternatives, num_conditions)

    if not result["success"]:
        return result

    return {
        "success": True,
        "alternatives_names": result["alternatives_names"],
        "conditions_names": result["conditions_names"],
        "cost_matrix": result["cost_matrix"],
        "error": None,
    }


def extract_savage_matrix(data, num_alternatives, num_conditions):
    """
    Extract Savage matrix from table data

    Args:
        data: Table data (list of rows)
        num_alternatives: Number of alternatives
        num_conditions: Number of conditions

    Returns:
        dict with alternatives names, conditions names, and cost matrix
    """
    if len(data) < 2:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": "Insufficient data for Savage matrix",
        }

    # First row should contain conditions names
    conditions_row = data[0]
    conditions_names = []
    for j in range(1, len(conditions_row)):  # Skip first column (alternatives)
        cell_value = str(conditions_row[j]).strip()
        if cell_value and cell_value.lower() != "nan":
            conditions_names.append(cell_value)
        else:
            conditions_names.append(f"Condition {j}")

    # Validate conditions count
    if len(conditions_names) != num_conditions:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Invalid number of conditions. Expected {num_conditions}, found {len(conditions_names)}",
        }

    # Extract alternatives names and cost matrix
    alternatives_names = []
    cost_matrix = []

    for i in range(1, len(data)):  # Skip header row
        row = data[i]
        if not row or len(row) == 0:
            continue

        # First column should contain alternative name
        alt_name = str(row[0]).strip()
        if alt_name and alt_name.lower() != "nan":
            alternatives_names.append(alt_name)
        else:
            alternatives_names.append(f"Alternative {i}")

        # Extract cost values
        cost_row = []
        for j in range(1, len(row)):  # Skip first column
            cell_value = str(row[j]).strip()
            if cell_value and cell_value.lower() != "nan":
                try:
                    # Handle checkmarks and other symbols - extract only numbers
                    import re

                    numbers = re.findall(r"-?\d+\.?\d*", cell_value)
                    if numbers:
                        value = float(numbers[0])
                        cost_row.append(value)
                    else:
                        return {
                            "success": False,
                            "alternatives_names": [],
                            "conditions_names": [],
                            "cost_matrix": [],
                            "error": f"Invalid cost value '{cell_value}' at position ({i},{j})",
                        }
                except (ValueError, TypeError):
                    return {
                        "success": False,
                        "alternatives_names": [],
                        "conditions_names": [],
                        "cost_matrix": [],
                        "error": f"Invalid cost value '{cell_value}' at position ({i},{j})",
                    }
            else:
                cost_row.append(0.0)

        # Pad row if necessary
        while len(cost_row) < num_conditions:
            cost_row.append(0.0)

        cost_matrix.append(cost_row)

    # Validate alternatives count
    if len(alternatives_names) != num_alternatives:
        return {
            "success": False,
            "alternatives_names": [],
            "conditions_names": [],
            "cost_matrix": [],
            "error": f"Invalid number of alternatives. Expected {num_alternatives}, found {len(alternatives_names)}",
        }

    return {
        "success": True,
        "alternatives_names": alternatives_names,
        "conditions_names": conditions_names,
        "cost_matrix": cost_matrix,
        "error": None,
    }


def process_experts_file(file, num_experts, num_alternatives, upload_folder="uploads"):
    """
    Process uploaded file specifically for experts evaluation analysis
    Handles files with two tables: competency weights and evaluation matrix

    Args:
        file: Flask file object
        num_experts: Number of experts
        num_alternatives: Number of alternatives
        upload_folder: Folder to save uploaded files

    Returns:
        dict: {'success': bool, 'competency_matrix': list, 'evaluation_matrix': list,
               'alternatives_names': list, 'error': str}
    """
    try:
        # Validate file
        is_valid, error = validate_file(file)
        if not is_valid:
            return {
                "success": False,
                "competency_matrix": [],
                "evaluation_matrix": [],
                "alternatives_names": [],
                "error": error,
            }

        # Save file
        file_path = save_uploaded_file(file, upload_folder)

        try:
            # Parse file based on extension
            file_ext = os.path.splitext(file.filename.lower())[1]

            if file_ext in [".xlsx", ".xls"]:
                data = parse_excel_data(file_path)
            elif file_ext == ".csv":
                data = parse_csv_data_with_empty_rows(file_path)
            else:
                return {
                    "success": False,
                    "competency_matrix": [],
                    "evaluation_matrix": [],
                    "alternatives_names": [],
                    "error": "Unsupported file format",
                }

            # Process experts data
            result = extract_experts_data(data, num_experts, num_alternatives)

            # Clean up file
            cleanup_file(file_path)

            return result

        except Exception as e:
            # Clean up file on error
            cleanup_file(file_path)
            return {
                "success": False,
                "competency_matrix": [],
                "evaluation_matrix": [],
                "alternatives_names": [],
                "error": f"Error processing file: {str(e)}",
            }

    except Exception as e:
        return {
            "success": False,
            "competency_matrix": [],
            "evaluation_matrix": [],
            "alternatives_names": [],
            "error": f"Upload error: {str(e)}",
        }


def extract_experts_data(data, num_experts, num_alternatives):
    """
    Extract experts evaluation data from file

    For experts evaluation analysis:
    - First table: competency matrix (expert weights for each criterion)
    - Second table: evaluation matrix (expert ratings for each alternative)
    - Alternative names are taken from column headers of the second table

    Args:
        data: Raw file data
        num_experts: Number of experts
        num_alternatives: Number of alternatives

    Returns:
        dict with competency matrix, evaluation matrix, and alternative names
    """
    try:
        if not data or len(data) == 0:
            return {
                "success": False,
                "competency_matrix": [],
                "evaluation_matrix": [],
                "alternatives_names": [],
                "error": "File is empty",
            }

        # Clean data - remove completely empty rows
        data = [row for row in data if any(str(cell).strip() for cell in row)]

        if len(data) == 0:
            return {
                "success": False,
                "competency_matrix": [],
                "evaluation_matrix": [],
                "alternatives_names": [],
                "error": "No data found in file",
            }

        # Find table boundaries by looking for empty or mostly empty rows
        table_boundaries = []
        current_start = 0

        for i, row in enumerate(data):
            # Check if row is empty or mostly empty (less than 2 non-empty cells)
            non_empty_cells = sum(
                1
                for cell in row
                if str(cell).strip() and str(cell).strip().lower() != "nan"
            )
            if non_empty_cells < 2:  # Row is empty or has only 1 non-empty cell
                if i > current_start:
                    table_boundaries.append((current_start, i))
                current_start = i + 1

        # Add the last table if it doesn't end with empty row
        if current_start < len(data):
            table_boundaries.append((current_start, len(data)))

        # If no empty rows found, try to split based on data patterns
        if not table_boundaries or len(table_boundaries) < 2:
            # Look for rows that might be headers (contain "Експерти" and multiple criteria)
            header_rows = []
            for i, row in enumerate(data):
                row_text = " ".join(
                    str(cell).strip().lower() for cell in row if str(cell).strip()
                )
                # Look for rows that contain "експерти" and have multiple criteria (not just expert names)
                if (
                    "експерти" in row_text
                    and len(row_text.split())
                    > 4  # More than just "Експерти" + 3 criteria
                    and not any(
                        name in row_text
                        for name in ["експерт 1", "експерт 2", "експерт 3", "експерт 4"]
                    )
                ):
                    header_rows.append(i)

            # If we found at least 2 potential header rows, split there
            if len(header_rows) >= 2:
                table_boundaries = []
                for i in range(len(header_rows)):
                    start = header_rows[i]
                    end = header_rows[i + 1] if i + 1 < len(header_rows) else len(data)
                    table_boundaries.append((start, end))
            else:
                # Fallback: treat entire data as one table
                table_boundaries = [(0, len(data))]

        if len(table_boundaries) < 2:
            return {
                "success": False,
                "competency_matrix": [],
                "evaluation_matrix": [],
                "alternatives_names": [],
                "error": "File must contain at least 2 tables (competency and evaluation)",
            }

        # Process first table (competency matrix)
        competency_data = data[table_boundaries[0][0] : table_boundaries[0][1]]
        competency_result = extract_competency_matrix(competency_data, num_experts)

        if not competency_result["success"]:
            return {
                "success": False,
                "competency_matrix": [],
                "evaluation_matrix": [],
                "alternatives_names": [],
                "error": f"Competency matrix error: {competency_result['error']}",
            }

        # Process second table (evaluation matrix)
        evaluation_data = data[table_boundaries[1][0] : table_boundaries[1][1]]
        evaluation_result = extract_evaluation_matrix(
            evaluation_data, num_experts, num_alternatives
        )

        if not evaluation_result["success"]:
            return {
                "success": False,
                "competency_matrix": [],
                "evaluation_matrix": [],
                "alternatives_names": [],
                "error": f"Evaluation matrix error: {evaluation_result['error']}",
            }

        return {
            "success": True,
            "competency_matrix": competency_result["matrix"],
            "evaluation_matrix": evaluation_result["matrix"],
            "alternatives_names": evaluation_result["alternatives_names"],
            "error": None,
        }

    except Exception as e:
        return {
            "success": False,
            "competency_matrix": [],
            "evaluation_matrix": [],
            "alternatives_names": [],
            "error": f"Error processing experts data: {str(e)}",
        }


def extract_competency_matrix(data, num_experts):
    """
    Extract competency matrix from first table

    Args:
        data: Table data (list of rows)
        num_experts: Number of experts

    Returns:
        dict with competency matrix
    """
    try:
        if len(data) < num_experts + 1:  # +1 for header
            return {
                "success": False,
                "matrix": [],
                "error": f"Insufficient data for competency matrix. Expected at least {num_experts + 1} rows",
            }

        # Extract competency matrix (skip first row which contains headers)
        matrix = []
        for i in range(1, min(num_experts + 1, len(data))):
            row = []
            for j in range(1, min(6, len(data[i]))):  # 5 criteria + 1 for expert name
                cell_value = str(data[i][j]).strip()
                try:
                    value = float(cell_value)
                    row.append(value)
                except (ValueError, TypeError):
                    return {
                        "success": False,
                        "matrix": [],
                        "error": f"Invalid competency value '{cell_value}' at position ({i},{j})",
                    }

            # Pad row if necessary
            while len(row) < 5:
                row.append(0.0)
            matrix.append(row)

        # Pad matrix if necessary
        while len(matrix) < num_experts:
            matrix.append([0.0] * 5)

        return {
            "success": True,
            "matrix": matrix,
            "error": None,
        }

    except Exception as e:
        return {
            "success": False,
            "matrix": [],
            "error": f"Error processing competency matrix: {str(e)}",
        }


def extract_evaluation_matrix(data, num_experts, num_alternatives):
    """
    Extract evaluation matrix from second table

    Args:
        data: Table data (list of rows)
        num_experts: Number of experts
        num_alternatives: Number of alternatives

    Returns:
        dict with evaluation matrix and alternative names
    """
    try:
        if len(data) < num_experts + 1:  # +1 for header
            return {
                "success": False,
                "matrix": [],
                "alternatives_names": [],
                "error": f"Insufficient data for evaluation matrix. Expected at least {num_experts + 1} rows",
            }

        # Extract alternative names from header row (excluding first column)
        alternatives_names = []
        for j in range(1, min(num_alternatives + 1, len(data[0]))):
            cell_value = str(data[0][j]).strip()
            if cell_value and cell_value != "":
                alternatives_names.append(cell_value)

        if len(alternatives_names) != num_alternatives:
            return {
                "success": False,
                "matrix": [],
                "alternatives_names": [],
                "error": f"Expected {num_alternatives} alternative names, found {len(alternatives_names)}",
            }

        # Extract evaluation matrix (skip first row which contains headers)
        matrix = []
        for i in range(1, min(num_experts + 1, len(data))):
            row = []
            for j in range(1, min(num_alternatives + 1, len(data[i]))):
                cell_value = str(data[i][j]).strip()
                try:
                    # Handle checkmarks and other symbols - extract only numbers
                    import re

                    numbers = re.findall(r"-?\d+\.?\d*", cell_value)
                    if numbers:
                        value = float(numbers[0])
                    else:
                        return {
                            "success": False,
                            "matrix": [],
                            "alternatives_names": [],
                            "error": f"Invalid evaluation value '{cell_value}' at position ({i},{j})",
                        }
                    row.append(value)
                except (ValueError, TypeError):
                    return {
                        "success": False,
                        "matrix": [],
                        "alternatives_names": [],
                        "error": f"Invalid evaluation value '{cell_value}' at position ({i},{j})",
                    }

            # Pad row if necessary
            while len(row) < num_alternatives:
                row.append(0.0)
            matrix.append(row)

        # Pad matrix if necessary
        while len(matrix) < num_experts:
            matrix.append([0.0] * num_alternatives)

        return {
            "success": True,
            "matrix": matrix,
            "alternatives_names": alternatives_names,
            "error": None,
        }

    except Exception as e:
        return {
            "success": False,
            "matrix": [],
            "alternatives_names": [],
            "error": f"Error processing evaluation matrix: {str(e)}",
        }
