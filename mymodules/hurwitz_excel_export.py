#!/usr/bin/env python3
"""
Excel export functionality for Hurwitz method
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
import io


class HurwitzExcelExporter:
    def __init__(self):
        self.workbook = Workbook()

        # Define colors matching AHP style
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        self.subheader_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        self.data_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        self.sum_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

        # Define fonts
        self.header_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        self.subheader_font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        self.data_font = Font(name="Arial", size=11, color="000000")
        self.sum_font = Font(name="Arial", size=11, bold=True, color="000000")

        # Define borders
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_header_style(self, cell, text=None, color="366092"):
        """Set header cell style"""
        if text is not None:
            cell.value = text
        cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border

    def set_subheader_style(self, cell, text=None, color="4F81BD"):
        """Set subheader cell style"""
        if text is not None:
            cell.value = text
        cell.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border

    def set_data_style(self, cell, value=None, bold=False):
        """Set data cell style"""
        if value is not None:
            cell.value = value
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border

    def set_number_style(self, cell, value, decimal_places=4, bold=False):
        """Set number cell style with specific decimal places"""
        if isinstance(value, (int, float)):
            cell.value = round(value, decimal_places)
        else:
            cell.value = value
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border
        if isinstance(value, (int, float)):
            if decimal_places == 0:
                cell.number_format = "0"
            else:
                cell.number_format = "0.000"

    def set_sum_style(self, cell, value=None, bold=True):
        """Set sum cell style"""
        if value is not None:
            cell.value = value
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border

    def set_row_height_for_text(self, ws, row, text, max_width=50):
        """Dynamically set row height based on text length"""
        if text and len(str(text)) > max_width:
            # Calculate approximate lines needed
            lines = len(str(text)) // max_width + 1
            # Set row height (default is 15, increase for more lines)
            ws.row_dimensions[row].height = max(15, lines * 15)

    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                try:
                    if hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    else:
                        # Find a non-merged cell for this column
                        for row_cell in ws[cell.coordinate[0]]:
                            if hasattr(row_cell, "column_letter"):
                                column_letter = row_cell.column_letter
                                break
                        if not column_letter:
                            continue

                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue

            if column_letter:
                adjusted_width = min(max(max_length + 2, 15), 60)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_general_info_sheet(self, analysis_data):
        """Create general information sheet"""
        ws = self.workbook.create_sheet("Загальна інформація")

        # Header
        self.set_header_style(ws["A1"], "Звіт аналізу Критерію Гурвіца")
        ws.merge_cells("A1:D1")

        # Method info
        self.set_data_style(ws["A3"], "Метод:")
        self.set_data_style(ws["B3"], "Критерій Гурвіца")
        self.set_data_style(ws["A4"], "ID аналізу:")
        self.set_data_style(ws["B4"], analysis_data.get("method_id", "N/A"))
        self.set_data_style(ws["A5"], "Коефіцієнт α:")
        self.set_data_style(ws["B5"], analysis_data.get("alpha", "N/A"))

        # Task description
        task_description = analysis_data.get("hurwitz_task", "Немає опису завдання")
        self.set_data_style(ws["A7"], "Опис завдання:")
        self.set_data_style(ws["A8"], task_description)
        ws.merge_cells("A8:D8")

        # Set text wrapping for task description
        ws["A8"].alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        self.set_row_height_for_text(ws, 8, task_description)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_cost_matrix_sheet(self, analysis_data):
        """Create cost matrix sheet"""
        ws = self.workbook.create_sheet("Матриця витрат")

        # Header
        self.set_header_style(ws["A1"], "Матриця витрат")
        ws.merge_cells("A1:E1")

        name_alternatives = analysis_data.get("name_alternatives", [])
        name_conditions = analysis_data.get("name_conditions", [])
        cost_matrix = analysis_data.get("cost_matrix", [])

        if not name_alternatives or not name_conditions or not cost_matrix:
            self.set_data_style(ws["A3"], "Немає даних для матриці витрат")
            return

        # Row headers
        self.set_subheader_style(ws["A3"], "Альтернативи")

        # Column headers
        for i, condition in enumerate(name_conditions):
            col_letter = chr(66 + i)  # B, C, D, etc.
            self.set_subheader_style(ws[f"{col_letter}3"], condition)

        # Fill matrix data
        for i, alternative in enumerate(name_alternatives):
            row = i + 4
            self.set_data_style(ws[f"A{row}"], alternative)

            for j, value in enumerate(cost_matrix[i]):
                col_letter = chr(66 + j)
                try:
                    int_value = int(value)
                    self.set_number_style(ws[f"{col_letter}{row}"], int_value, 0)
                except (ValueError, TypeError):
                    self.set_data_style(ws[f"{col_letter}{row}"], value)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_hurwitz_values_sheet(self, analysis_data):
        """Create Hurwitz values sheet"""
        ws = self.workbook.create_sheet("Значення Гурвіца")

        # Header
        self.set_header_style(ws["A1"], "Значення Гурвіца")
        ws.merge_cells("A1:F1")

        name_alternatives = analysis_data.get("name_alternatives", [])
        min_values = analysis_data.get("min_values", [])
        max_values = analysis_data.get("max_values", [])
        hurwitz_values = analysis_data.get("hurwitz_values", [])
        alpha = analysis_data.get("alpha", 0.5)

        if (
            not name_alternatives
            or not min_values
            or not max_values
            or not hurwitz_values
        ):
            self.set_data_style(ws["A3"], "Немає даних для значень Гурвіца")
            return

        # Row headers
        self.set_subheader_style(ws["A3"], "Альтернативи")
        self.set_subheader_style(ws["B3"], "Мін")
        self.set_subheader_style(ws["C3"], "Макс")
        self.set_subheader_style(ws["D3"], f"α = {alpha}")
        self.set_subheader_style(ws["E3"], "Гурвіца")

        # Fill data
        for i, alternative in enumerate(name_alternatives):
            row = i + 4
            self.set_data_style(ws[f"A{row}"], alternative)
            self.set_number_style(ws[f"B{row}"], int(min_values[i]), 0)
            self.set_number_style(ws[f"C{row}"], int(max_values[i]), 0)
            self.set_number_style(ws[f"D{row}"], alpha, 3)
            self.set_number_style(ws[f"E{row}"], hurwitz_values[i], 3)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_optimal_variants_sheet(self, analysis_data):
        """Create optimal variants sheet"""
        ws = self.workbook.create_sheet("Оптимальні варіанти")

        # Header
        self.set_header_style(ws["A1"], "Оптимальні варіанти")
        ws.merge_cells("A1:B1")

        # Table headers
        self.set_subheader_style(ws["A3"], "Альтернативи")
        self.set_subheader_style(ws["B3"], "Значення Гурвіца")

        name_alternatives = analysis_data.get("name_alternatives", [])
        hurwitz_values = analysis_data.get("hurwitz_values", [])

        if not name_alternatives or not hurwitz_values:
            self.set_data_style(ws["A5"], "Немає даних для оптимальних варіантів")
            return

        # Create ranking data
        ranking_data = []
        for i, (alternative, hurwitz_value) in enumerate(
            zip(name_alternatives, hurwitz_values)
        ):
            ranking_data.append((alternative, hurwitz_value))

        # Sort by Hurwitz value in descending order (higher is better)
        ranking_data.sort(key=lambda x: x[1], reverse=True)

        # Fill ranking table
        for i, (alternative, hurwitz_value) in enumerate(ranking_data):
            row = i + 5
            self.set_data_style(ws[f"A{row}"], alternative)
            self.set_number_style(ws[f"B{row}"], hurwitz_value, 3)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_ranking_chart(self, analysis_data):
        """Create ranking chart"""
        ws = self.workbook.create_sheet("Графіки")

        # Header
        self.set_header_style(ws["A1"], "Графік значень Гурвіца")
        ws.merge_cells("A1:D1")

        name_alternatives = analysis_data.get("name_alternatives", [])
        hurwitz_values = analysis_data.get("hurwitz_values", [])

        if not name_alternatives or not hurwitz_values:
            self.set_data_style(ws["A3"], "Немає даних для графіка")
            return

        # Prepare data for chart
        self.set_subheader_style(ws["A3"], "Альтернативи")
        self.set_subheader_style(ws["B3"], "Значення Гурвіца")

        # Fill data in original order (no sorting)
        for i, (alternative, hurwitz_value) in enumerate(
            zip(name_alternatives, hurwitz_values)
        ):
            row = i + 4
            self.set_data_style(ws[f"A{row}"], alternative)
            self.set_number_style(ws[f"B{row}"], hurwitz_value, 3)

        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Значення Гурвіца альтернатив"
        chart.y_axis.title = "Значення"
        chart.x_axis.title = "Альтернативи"

        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(name_alternatives))
        cats = Reference(
            ws, min_col=1, min_row=4, max_row=4 + len(name_alternatives) - 1
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Position chart
        ws.add_chart(chart, "A10")

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_conclusion_sheet(self, analysis_data):
        """Create conclusion sheet"""
        ws = self.workbook.create_sheet("Висновки")

        # Header
        self.set_header_style(ws["A1"], "Висновки")
        ws.merge_cells("A1:D1")

        # Conclusion
        optimal_message = analysis_data.get("optimal_message", "Аналіз завершено")
        self.set_data_style(ws["A3"], "Результат аналізу:")
        self.set_data_style(ws["A4"], optimal_message)
        ws.merge_cells("A4:D4")

        # Set text wrapping for conclusion
        ws["A4"].alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        self.set_row_height_for_text(ws, 4, optimal_message)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def generate_hurwitz_analysis_excel(self, analysis_data):
        """Generate complete hurwitz analysis Excel file"""
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        # Create all sheets
        self.create_general_info_sheet(analysis_data)
        self.create_cost_matrix_sheet(analysis_data)
        self.create_hurwitz_values_sheet(analysis_data)
        self.create_optimal_variants_sheet(analysis_data)
        self.create_ranking_chart(analysis_data)
        self.create_conclusion_sheet(analysis_data)

        return self.workbook

    def save_to_bytes(self):
        """Save workbook to bytes"""
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
