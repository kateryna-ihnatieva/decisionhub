#!/usr/bin/env python3
"""
Excel export functionality for Laplasa method
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
import io
from datetime import datetime


class LaplasaExcelExporter:
    def __init__(self):
        self.workbook = Workbook()

        # Define colors matching AHP style
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        self.subheader_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        self.data_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        self.sum_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

        # Define fonts
        self.header_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        self.subheader_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        self.data_font = Font(name="Calibri", size=11, color="000000")
        self.sum_font = Font(name="Calibri", size=11, bold=True, color="000000")

        # Define borders
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_header_style(self, cell):
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.border = self.thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_subheader_style(self, cell):
        cell.fill = self.subheader_fill
        cell.font = self.subheader_font
        cell.border = self.thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_data_style(self, cell):
        cell.fill = self.data_fill
        cell.font = self.data_font
        cell.border = self.thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_sum_style(self, cell):
        cell.fill = self.sum_fill
        cell.font = self.sum_font
        cell.border = self.thin_border
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def set_number_style(self, cell, value):
        if isinstance(value, (int, float)):
            cell.value = int(value) if value == int(value) else round(value, 3)
        else:
            cell.value = value
        cell.number_format = "0.000" if isinstance(value, float) else "0"
        self.set_data_style(cell)

    def set_row_height_for_text(self, ws, row, text, min_height=30):
        """Set row height based on text length for better wrapping"""
        if text and len(str(text)) > 50:  # For long text
            # Calculate approximate height needed
            text_length = len(str(text))
            lines_needed = max(1, text_length // 50)  # Approximate lines
            height = max(min_height, lines_needed * 15)  # 15 points per line
            ws.row_dimensions[row].height = height

    def create_general_info_sheet(self, analysis_data):
        ws = self.workbook.active
        ws.title = "Загальна інформація"

        # Header
        ws["A1"] = "Звіт аналізу Критерію Лапласа"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:D1")

        # Method info
        ws["A3"] = "Метод аналізу:"
        ws["B3"] = "Критерій Лапласа"
        self.set_data_style(ws["A3"])
        self.set_data_style(ws["B3"])

        ws["A4"] = "ID аналізу:"
        ws["B4"] = analysis_data.get("method_id", "N/A")
        self.set_data_style(ws["A4"])
        self.set_data_style(ws["B4"])

        ws["A5"] = "Дата створення:"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.set_data_style(ws["A5"])
        self.set_data_style(ws["B5"])

        ws["A6"] = "Опис завдання:"
        self.set_data_style(ws["A6"])

        task_description = analysis_data.get("laplasa_task", "Аналіз критерію Лапласа")
        ws["A7"] = task_description
        self.set_data_style(ws["A7"])
        ws.merge_cells("A7:D7")
        self.set_row_height_for_text(ws, 7, task_description)

        # Matrix type
        matrix_type = analysis_data.get("matrix_type", "profit")
        matrix_type_text = "Прибуток" if matrix_type == "profit" else "Затрати"
        ws["A8"] = "Тип матриці:"
        ws["B8"] = matrix_type_text
        self.set_data_style(ws["A8"])
        self.set_data_style(ws["B8"])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_cost_matrix_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Матриця витрат")

        # Header
        ws["A1"] = "Матриця витрат"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:D1")

        name_alternatives = analysis_data.get("name_alternatives", [])
        name_conditions = analysis_data.get("name_conditions", [])
        cost_matrix = analysis_data.get("cost_matrix", [])

        if not name_alternatives or not name_conditions or not cost_matrix:
            ws["A3"] = "Немає даних для матриці витрат"
            self.set_data_style(ws["A3"])
            return

        # Table headers
        ws["A3"] = "Альтернативи"
        self.set_subheader_style(ws["A3"])

        # Condition headers
        for i, condition in enumerate(name_conditions):
            col_letter = chr(66 + i)  # B, C, D, etc.
            ws[f"{col_letter}3"] = condition
            self.set_subheader_style(ws[f"{col_letter}3"])

        # M column
        m_col = chr(66 + len(name_conditions))
        ws[f"{m_col}3"] = "M"
        self.set_subheader_style(ws[f"{m_col}3"])

        # Fill matrix data
        for i, alternative in enumerate(name_alternatives):
            row = i + 4
            ws[f"A{row}"] = alternative
            self.set_data_style(ws[f"A{row}"])

            # Fill cost matrix values
            for j, condition in enumerate(name_conditions):
                col_letter = chr(66 + j)  # B, C, D, etc.
                if i < len(cost_matrix) and j < len(cost_matrix[i]):
                    value = cost_matrix[i][j]
                    try:
                        # Convert to int if possible
                        int_value = int(value)
                        ws[f"{col_letter}{row}"] = int_value
                        self.set_number_style(ws[f"{col_letter}{row}"], int_value)
                    except (ValueError, TypeError):
                        ws[f"{col_letter}{row}"] = value
                        self.set_data_style(ws[f"{col_letter}{row}"])

        # Add M values (optimal variants)
        optimal_variants = analysis_data.get("optimal_variants", [])
        for i in range(len(name_alternatives)):
            row = i + 4
            m_col = chr(66 + len(name_conditions))

            # Use optimal variant value
            if i < len(optimal_variants):
                m_value = optimal_variants[i]
                ws[f"{m_col}{row}"] = round(m_value, 3)
                self.set_number_style(ws[f"{m_col}{row}"], m_value)
            else:
                ws[f"{m_col}{row}"] = 0
                self.set_number_style(ws[f"{m_col}{row}"], 0)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_optimal_variants_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Оптимальні варіанти")

        # Header
        ws["A1"] = "Оптимальні варіанти"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:B1")

        # Table headers
        matrix_type = analysis_data.get("matrix_type", "profit")
        header_text = (
            "Очікувана вигода" if matrix_type == "profit" else "Очікувані затрати"
        )
        ws["A3"] = "Альтернативи"
        ws["B3"] = header_text
        self.set_subheader_style(ws["A3"])
        self.set_subheader_style(ws["B3"])

        name_alternatives = analysis_data.get("name_alternatives", [])
        optimal_variants = analysis_data.get("optimal_variants", [])

        if not name_alternatives or not optimal_variants:
            ws["A5"] = "Немає даних для оптимальних варіантів"
            self.set_data_style(ws["A5"])
            return

        # Create ranking data
        matrix_type = analysis_data.get("matrix_type", "profit")
        ranking_data = []
        for i, (alternative, variant) in enumerate(
            zip(name_alternatives, optimal_variants)
        ):
            ranking_data.append((alternative, variant))

        # Sort by variant value: descending for profit, ascending for cost
        ranking_data.sort(key=lambda x: x[1], reverse=(matrix_type == "profit"))

        # Fill ranking table
        for i, (alternative, variant) in enumerate(ranking_data):
            row = i + 5
            ws[f"A{row}"] = alternative
            ws[f"B{row}"] = round(variant, 3)

            self.set_data_style(ws[f"A{row}"])
            self.set_number_style(ws[f"B{row}"], variant)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_chart_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Графіки")

        # Header
        ws["A1"] = "Графік оптимальних варіантів"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:D1")

        name_alternatives = analysis_data.get("name_alternatives", [])
        optimal_variants = analysis_data.get("optimal_variants", [])

        if not name_alternatives or not optimal_variants:
            ws["A3"] = "Немає даних для графіка"
            self.set_data_style(ws["A3"])
            return

        # Add data for chart
        matrix_type = analysis_data.get("matrix_type", "profit")
        header_text = (
            "Очікувана вигода" if matrix_type == "profit" else "Очікувані затрати"
        )
        ws["A3"] = "Альтернативи"
        ws["B3"] = header_text
        self.set_subheader_style(ws["A3"])
        self.set_subheader_style(ws["B3"])

        for i, (alternative, variant) in enumerate(
            zip(name_alternatives, optimal_variants)
        ):
            row = i + 4
            ws[f"A{row}"] = alternative
            ws[f"B{row}"] = round(variant, 3)
            self.set_data_style(ws[f"A{row}"])
            self.set_number_style(ws[f"B{row}"], variant)

        # Create chart
        matrix_type = analysis_data.get("matrix_type", "profit")
        y_axis_title = (
            "Очікувана вигода" if matrix_type == "profit" else "Очікувані затрати"
        )
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Оптимальні варіанти"
        chart.y_axis.title = y_axis_title
        chart.x_axis.title = "Альтернативи"

        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(name_alternatives))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(name_alternatives))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Position chart
        ws.add_chart(chart, "E3")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_conclusion_sheet(self, analysis_data):
        ws = self.workbook.create_sheet("Висновки")

        # Header
        ws["A1"] = "Висновки"
        self.set_header_style(ws["A1"])
        ws.merge_cells("A1:D1")

        optimal_message = analysis_data.get("optimal_message", "Аналіз завершено")

        ws["A3"] = "Результат аналізу:"
        self.set_subheader_style(ws["A3"])

        ws["A4"] = optimal_message
        self.set_data_style(ws["A4"])
        ws.merge_cells("A4:D4")
        self.set_row_height_for_text(ws, 4, optimal_message)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if column_letter is None and hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    def generate_laplasa_analysis_excel(self, analysis_data):
        """Generate complete Excel file for Laplasa analysis"""

        # Create sheets in order
        self.create_general_info_sheet(analysis_data)
        self.create_cost_matrix_sheet(analysis_data)
        self.create_optimal_variants_sheet(analysis_data)
        self.create_chart_sheet(analysis_data)
        self.create_conclusion_sheet(analysis_data)

        return self.workbook

    def save_to_bytes(self):
        """Save workbook to bytes"""
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
