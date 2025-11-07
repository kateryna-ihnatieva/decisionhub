#!/usr/bin/env python3
"""
Excel export functionality for Savage method
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
import io


class SavageExcelExporter:
    def __init__(self):
        self.workbook = Workbook()

        # Define colors matching AHP style
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        self.subheader_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        self.data_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        self.sum_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

        # Define fonts
        self.header_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        self.subheader_font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        self.data_font = Font(name="Arial", size=11, color="000000")
        self.sum_font = Font(name="Arial", size=11, bold=True, color="000000")

        # Define borders
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def set_header_style(self, cell, text=None, color="366092"):
        """Set header cell style"""
        if text is not None:
            cell.value = text
        cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border

    def set_subheader_style(self, cell, text=None, color="4F81BD"):
        """Set subheader cell style"""
        if text is not None:
            cell.value = text
        cell.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border

    def set_data_style(self, cell, value=None, bold=False):
        """Set data cell style"""
        if value is not None:
            cell.value = value
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border

    def set_number_style(self, cell, value, decimal_places=4, bold=False):
        """Set number cell style with specific decimal places"""
        if isinstance(value, (int, float)):
            cell.value = round(value, decimal_places)
        else:
            cell.value = value
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border
        if isinstance(value, (int, float)):
            if decimal_places == 0:
                cell.number_format = "0"
            else:
                cell.number_format = "0.000"

    def set_sum_style(self, cell, value=None, bold=True):
        """Set sum cell style"""
        if value is not None:
            cell.value = value
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = self.thin_border

    def set_row_height_for_text(self, ws, row, text, max_width=50):
        """Dynamically set row height based on text length"""
        if text and len(str(text)) > max_width:
            # Calculate approximate lines needed
            lines = len(str(text)) // max_width + 1
            # Set row height (default is 15, increase for more lines)
            ws.row_dimensions[row].height = max(15, lines * 15)

    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                try:
                    if hasattr(cell, "column_letter"):
                        column_letter = cell.column_letter
                    else:
                        # Find a non-merged cell for this column
                        for row_cell in ws[cell.coordinate[0]]:
                            if hasattr(row_cell, "column_letter"):
                                column_letter = row_cell.column_letter
                                break
                        if not column_letter:
                            continue

                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue

            if column_letter:
                adjusted_width = min(max(max_length + 2, 15), 60)
                ws.column_dimensions[column_letter].width = adjusted_width

    def create_general_info_sheet(self, analysis_data):
        """Create general information sheet"""
        ws = self.workbook.create_sheet("Загальна інформація")

        # Header
        self.set_header_style(ws["A1"], "Звіт аналізу Критерію Севіджа")
        ws.merge_cells("A1:D1")

        # Method info
        self.set_data_style(ws["A3"], "Метод:")
        self.set_data_style(ws["B3"], "Критерій Севіджа")
        self.set_data_style(ws["A4"], "ID аналізу:")
        self.set_data_style(ws["B4"], analysis_data.get("method_id", "N/A"))

        # Task description
        task_description = analysis_data.get("savage_task", "Немає опису завдання")
        self.set_data_style(ws["A6"], "Опис завдання:")
        self.set_data_style(ws["A7"], task_description)
        ws.merge_cells("A7:D7")

        # Set text wrapping for task description
        ws["A7"].alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        self.set_row_height_for_text(ws, 7, task_description)

        # Matrix type
        matrix_type = analysis_data.get("matrix_type", "profit")
        matrix_type_text = "Прибуток" if matrix_type == "profit" else "Затрати"
        self.set_data_style(ws["A8"], "Тип матриці:")
        self.set_data_style(ws["B8"], matrix_type_text)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_cost_matrix_sheet(self, analysis_data):
        """Create cost matrix sheet"""
        ws = self.workbook.create_sheet("Матриця витрат")

        # Header
        self.set_header_style(ws["A1"], "Матриця витрат")
        ws.merge_cells("A1:E1")

        name_alternatives = analysis_data.get("name_alternatives", [])
        name_conditions = analysis_data.get("name_conditions", [])
        cost_matrix = analysis_data.get("cost_matrix", [])

        if not name_alternatives or not name_conditions or not cost_matrix:
            self.set_data_style(ws["A3"], "Немає даних для матриці витрат")
            return

        # Row headers
        self.set_subheader_style(ws["A3"], "Альтернативи")

        # Column headers
        for i, condition in enumerate(name_conditions):
            col_letter = chr(66 + i)  # B, C, D, etc.
            self.set_subheader_style(ws[f"{col_letter}3"], condition)

        # Fill matrix data
        for i, alternative in enumerate(name_alternatives):
            row = i + 4
            self.set_data_style(ws[f"A{row}"], alternative)

            for j, value in enumerate(cost_matrix[i]):
                col_letter = chr(66 + j)
                try:
                    int_value = int(value)
                    self.set_number_style(ws[f"{col_letter}{row}"], int_value, 0)
                except (ValueError, TypeError):
                    self.set_data_style(ws[f"{col_letter}{row}"], value)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_loss_matrix_sheet(self, analysis_data):
        """Create loss matrix sheet"""
        ws = self.workbook.create_sheet("Матриця втрат")

        # Header
        self.set_header_style(ws["A1"], "Матриця втрат")
        ws.merge_cells("A1:E1")

        name_alternatives = analysis_data.get("name_alternatives", [])
        name_conditions = analysis_data.get("name_conditions", [])
        loss_matrix = analysis_data.get("loss_matrix", [])

        if not name_alternatives or not name_conditions or not loss_matrix:
            self.set_data_style(ws["A3"], "Немає даних для матриці втрат")
            return

        # Row headers
        self.set_subheader_style(ws["A3"], "Альтернативи")

        # Column headers
        for i, condition in enumerate(name_conditions):
            col_letter = chr(66 + i)  # B, C, D, etc.
            self.set_subheader_style(ws[f"{col_letter}3"], condition)

        # Max losses column
        max_col = chr(66 + len(name_conditions))
        self.set_subheader_style(ws[f"{max_col}3"], "Макс. втрати")

        # Fill matrix data
        for i, alternative in enumerate(name_alternatives):
            row = i + 4
            self.set_data_style(ws[f"A{row}"], alternative)

            for j, value in enumerate(loss_matrix[i]):
                col_letter = chr(66 + j)
                try:
                    int_value = int(value)
                    self.set_number_style(ws[f"{col_letter}{row}"], int_value, 0)
                except (ValueError, TypeError):
                    self.set_data_style(ws[f"{col_letter}{row}"], value)

        # Add max losses
        max_losses = analysis_data.get("max_losses", [])
        for i in range(len(name_alternatives)):
            row = i + 4
            max_col = chr(66 + len(name_conditions))

            # Use max loss value
            if i < len(max_losses):
                max_value = max_losses[i]
                self.set_number_style(ws[f"{max_col}{row}"], int(max_value), 0)
            else:
                self.set_number_style(ws[f"{max_col}{row}"], 0, 0)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_optimal_variants_sheet(self, analysis_data):
        """Create optimal variants sheet"""
        ws = self.workbook.create_sheet("Оптимальні варіанти")

        # Header
        self.set_header_style(ws["A1"], "Оптимальні варіанти")
        ws.merge_cells("A1:B1")

        # Table headers
        self.set_subheader_style(ws["A3"], "Альтернативи")
        self.set_subheader_style(ws["B3"], "Максимальні втрати")

        name_alternatives = analysis_data.get("name_alternatives", [])
        max_losses = analysis_data.get("max_losses", [])

        if not name_alternatives or not max_losses:
            self.set_data_style(ws["A5"], "Немає даних для оптимальних варіантів")
            return

        # Create ranking data
        ranking_data = []
        for i, (alternative, max_loss) in enumerate(zip(name_alternatives, max_losses)):
            ranking_data.append((alternative, max_loss))

        # Sort by max loss in ascending order (lower is better for Savage)
        ranking_data.sort(key=lambda x: x[1])

        # Fill ranking table
        for i, (alternative, max_loss) in enumerate(ranking_data):
            row = i + 5
            self.set_data_style(ws[f"A{row}"], alternative)
            self.set_number_style(ws[f"B{row}"], int(max_loss), 0)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_ranking_chart(self, analysis_data):
        """Create ranking chart"""
        ws = self.workbook.create_sheet("Графіки")

        # Header
        self.set_header_style(ws["A1"], "Графік максимальних втрат")
        ws.merge_cells("A1:D1")

        name_alternatives = analysis_data.get("name_alternatives", [])
        max_losses = analysis_data.get("max_losses", [])

        if not name_alternatives or not max_losses:
            self.set_data_style(ws["A3"], "Немає даних для графіка")
            return

        # Prepare data for chart
        self.set_subheader_style(ws["A3"], "Альтернативи")
        self.set_subheader_style(ws["B3"], "Максимальні втрати")

        # Fill data in original order (no sorting)
        for i, (alternative, max_loss) in enumerate(zip(name_alternatives, max_losses)):
            row = i + 4
            self.set_data_style(ws[f"A{row}"], alternative)
            self.set_number_style(ws[f"B{row}"], int(max_loss), 0)

        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Максимальні втрати альтернатив"
        chart.y_axis.title = "Втрати"
        chart.x_axis.title = "Альтернативи"

        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(name_alternatives))
        cats = Reference(
            ws, min_col=1, min_row=4, max_row=4 + len(name_alternatives) - 1
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Position chart
        ws.add_chart(chart, "A10")

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def create_conclusion_sheet(self, analysis_data):
        """Create conclusion sheet"""
        ws = self.workbook.create_sheet("Висновки")

        # Header
        self.set_header_style(ws["A1"], "Висновки")
        ws.merge_cells("A1:D1")

        # Conclusion
        optimal_message = analysis_data.get("optimal_message", "Аналіз завершено")
        self.set_data_style(ws["A3"], "Результат аналізу:")
        self.set_data_style(ws["A4"], optimal_message)
        ws.merge_cells("A4:D4")

        # Set text wrapping for conclusion
        ws["A4"].alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        self.set_row_height_for_text(ws, 4, optimal_message)

        # Auto-adjust column widths
        self.auto_adjust_columns(ws)

    def generate_savage_analysis_excel(self, analysis_data):
        """Generate complete savage analysis Excel file"""
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        # Create all sheets
        self.create_general_info_sheet(analysis_data)
        self.create_cost_matrix_sheet(analysis_data)
        self.create_loss_matrix_sheet(analysis_data)
        self.create_optimal_variants_sheet(analysis_data)
        self.create_ranking_chart(analysis_data)
        self.create_conclusion_sheet(analysis_data)

        return self.workbook

    def save_to_bytes(self):
        """Save workbook to bytes"""
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
