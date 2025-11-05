import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

# Create test data directory
os.makedirs("test_files", exist_ok=True)

# Common test scenario: Choosing the best smartphone
alternatives = ["iPhone 15", "Samsung Galaxy S24", "Google Pixel 8", "OnePlus 12"]
criteria = ["Price", "Camera Quality", "Battery Life", "Performance", "Design"]

print("Creating Excel test files for decision-making methods...")

# 1. HIERARCHY (AHP) TEST FILE
print("1. Creating Hierarchy (AHP) test file...")
wb_hierarchy = Workbook()
ws_hierarchy = wb_hierarchy.active
ws_hierarchy.title = "AHP Test Data"

# Criteria comparison matrix (5x5)
criteria_matrix = [
    [1, 3, 2, 4, 2],  # Price vs others
    [1 / 3, 1, 1 / 2, 2, 1 / 2],  # Camera vs others
    [1 / 2, 2, 1, 3, 1],  # Battery vs others
    [1 / 4, 1 / 2, 1 / 3, 1, 1 / 3],  # Performance vs others
    [1 / 2, 2, 1, 3, 1],  # Design vs others
]

# Write criteria matrix
ws_hierarchy["A1"] = "Criteria Comparison Matrix"
ws_hierarchy["A1"].font = Font(bold=True, size=14)

# Headers
for i, criterion in enumerate(criteria):
    ws_hierarchy.cell(row=2, column=i + 2, value=criterion).font = Font(bold=True)
    ws_hierarchy.cell(row=i + 3, column=1, value=criterion).font = Font(bold=True)

# Matrix data
for i in range(5):
    for j in range(5):
        ws_hierarchy.cell(row=i + 3, column=j + 2, value=criteria_matrix[i][j])

# Empty row separator
ws_hierarchy["A9"] = ""

# Alternatives comparison matrices for each criterion
for crit_idx, criterion in enumerate(criteria):
    start_row = 10 + crit_idx * 7

    ws_hierarchy.cell(row=start_row, column=1, value=f"{criterion} Comparison").font = (
        Font(bold=True, size=12)
    )

    # Headers
    for i, alt in enumerate(alternatives):
        ws_hierarchy.cell(row=start_row + 1, column=i + 2, value=alt).font = Font(
            bold=True
        )
        ws_hierarchy.cell(row=start_row + i + 2, column=1, value=alt).font = Font(
            bold=True
        )

    # Generate random but consistent comparison matrix
    np.random.seed(42 + crit_idx)  # Consistent random seed
    alt_matrix = np.ones((4, 4))
    for i in range(4):
        for j in range(i + 1, 4):
            # Generate random comparison value between 1/9 and 9
            val = np.random.choice(
                [
                    1 / 9,
                    1 / 8,
                    1 / 7,
                    1 / 6,
                    1 / 5,
                    1 / 4,
                    1 / 3,
                    1 / 2,
                    1,
                    2,
                    3,
                    4,
                    5,
                    6,
                    7,
                    8,
                    9,
                ]
            )
            alt_matrix[i][j] = val
            alt_matrix[j][i] = 1 / val

    # Write matrix
    for i in range(4):
        for j in range(4):
            ws_hierarchy.cell(
                row=start_row + i + 2, column=j + 2, value=round(alt_matrix[i][j], 3)
            )

wb_hierarchy.save("test_files/Hierarchy_Test_Data.xlsx")
print("   ✓ Hierarchy test file created")

# 2. BINARY RELATIONS TEST FILE
print("2. Creating Binary Relations test file...")
wb_binary = Workbook()
ws_binary = wb_binary.active
ws_binary.title = "Binary Relations Test Data"

ws_binary["A1"] = "Binary Relations Matrix (Smartphone Preferences)"
ws_binary["A1"].font = Font(bold=True, size=14)

# Headers
for i, alt in enumerate(alternatives):
    ws_binary.cell(row=2, column=i + 2, value=alt).font = Font(bold=True)
    ws_binary.cell(row=i + 3, column=1, value=alt).font = Font(bold=True)

# Binary relations matrix (-1, 0, 1)
# 1 = row is preferred to column, -1 = column is preferred to row, 0 = equal
binary_matrix = [
    [0, 1, 1, -1],  # iPhone vs others
    [-1, 0, 1, 1],  # Samsung vs others
    [-1, -1, 0, 1],  # Google vs others
    [1, -1, -1, 0],  # OnePlus vs others
]

for i in range(4):
    for j in range(4):
        ws_binary.cell(row=i + 3, column=j + 2, value=binary_matrix[i][j])

wb_binary.save("test_files/Binary_Relations_Test_Data.xlsx")
print("   ✓ Binary Relations test file created")

# 3. EXPERT EVALUATION TEST FILE
print("3. Creating Expert Evaluation test file...")
wb_experts = Workbook()
ws_experts = wb_experts.active
ws_experts.title = "Expert Evaluation Test Data"

# Competency matrix (4 experts x 5 criteria)
ws_experts["A1"] = "Expert Competency Matrix"
ws_experts["A1"].font = Font(bold=True, size=14)

# Headers
ws_experts.cell(row=2, column=1, value="Experts").font = Font(bold=True)
for i, criterion in enumerate(criteria):
    ws_experts.cell(row=2, column=i + 2, value=criterion).font = Font(bold=True)

# Expert names
expert_names = ["Expert 1", "Expert 2", "Expert 3", "Expert 4"]
for i, expert in enumerate(expert_names):
    ws_experts.cell(row=i + 3, column=1, value=expert).font = Font(bold=True)

# Competency scores (1-10 scale)
np.random.seed(123)
competency_matrix = np.random.randint(6, 11, (4, 5))  # Scores 6-10

for i in range(4):
    for j in range(5):
        ws_experts.cell(row=i + 3, column=j + 2, value=competency_matrix[i][j])

# Empty row separator
ws_experts["A8"] = ""

# Evaluation matrix (4 experts x 4 alternatives)
ws_experts["A9"] = "Expert Evaluation Matrix"
ws_experts["A9"].font = Font(bold=True, size=14)

# Headers
ws_experts.cell(row=10, column=1, value="Experts").font = Font(bold=True)
for i, alt in enumerate(alternatives):
    ws_experts.cell(row=10, column=i + 2, value=alt).font = Font(bold=True)

# Expert names
for i, expert in enumerate(expert_names):
    ws_experts.cell(row=i + 11, column=1, value=expert).font = Font(bold=True)

# Evaluation scores (1-10 scale)
np.random.seed(456)
evaluation_matrix = np.random.randint(5, 11, (4, 4))  # Scores 5-10

for i in range(4):
    for j in range(4):
        ws_experts.cell(row=i + 11, column=j + 2, value=evaluation_matrix[i][j])

wb_experts.save("test_files/Expert_Evaluation_Test_Data.xlsx")
print("   ✓ Expert Evaluation test file created")

# 4. DECISION METHODS TEST FILES (Laplace, Maximin, Savage, Hurwitz)
print("4. Creating Decision Methods test files...")

# Common cost/profit matrix for all decision methods
# Rows: Alternatives, Columns: Conditions (Market scenarios)
conditions = ["High Demand", "Medium Demand", "Low Demand", "Economic Crisis"]

# Generate realistic cost matrix (lower is better for cost, higher is better for profit)
np.random.seed(789)
base_costs = [800, 900, 700, 850]  # Base prices for each smartphone
cost_matrix = []

for i, base_cost in enumerate(base_costs):
    row = []
    for j in range(4):
        # Add market condition modifiers
        modifiers = [0.8, 1.0, 1.2, 1.5]  # High demand (cheaper), crisis (expensive)
        cost = int(base_cost * modifiers[j] + np.random.randint(-50, 51))
        row.append(max(100, cost))  # Ensure positive values
    cost_matrix.append(row)

# Create files for each decision method
methods = ["Laplace", "Maximin", "Savage", "Hurwitz"]

for method in methods:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{method} Test Data"

    ws["A1"] = f"{method} Method Test Data - Smartphone Selection"
    ws["A1"].font = Font(bold=True, size=14)

    # Headers
    ws.cell(row=2, column=1, value="Alternatives").font = Font(bold=True)
    for i, condition in enumerate(conditions):
        ws.cell(row=2, column=i + 2, value=condition).font = Font(bold=True)

    # Alternative names
    for i, alt in enumerate(alternatives):
        ws.cell(row=i + 3, column=1, value=alt).font = Font(bold=True)

    # Cost matrix data
    for i in range(4):
        for j in range(4):
            ws.cell(row=i + 3, column=j + 2, value=cost_matrix[i][j])

    wb.save(f"test_files/{method}_Test_Data.xlsx")
    print(f"   ✓ {method} test file created")

print("\n" + "=" * 60)
print("ALL TEST FILES CREATED SUCCESSFULLY!")
print("=" * 60)
print("\nFiles created in 'test_files/' directory:")
print("1. Hierarchy_Test_Data.xlsx - AHP method test data")
print("2. Binary_Relations_Test_Data.xlsx - Binary relations test data")
print("3. Expert_Evaluation_Test_Data.xlsx - Expert evaluation test data")
print("4. Laplace_Test_Data.xlsx - Laplace criterion test data")
print("5. Maximin_Test_Data.xlsx - Maximin criterion test data")
print("6. Savage_Test_Data.xlsx - Savage criterion test data")
print("7. Hurwitz_Test_Data.xlsx - Hurwitz criterion test data")
print("\nAll files use the same scenario: Choosing the best smartphone")
print("with 4 alternatives and consistent data structure for fair comparison.")
