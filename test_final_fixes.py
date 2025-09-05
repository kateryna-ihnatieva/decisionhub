#!/usr/bin/env python3
"""
Test script for final Experts Excel fixes
"""

import sys
import os

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from mymodules.experts_excel_export import ExpertsExcelExporter


def test_final_fixes():
    """Test final fixes for Experts Excel export"""

    print("🧪 Testing Final Experts Excel Fixes...")

    # Sample analysis data
    analysis_data = {
        "method_id": 123,
        "experts_task": "Тестова задача для експертних оцінок з довгим текстом",
        "table_competency": [
            [0.8, 0.7, 0.9, 0.6, 0.8],
            [0.9, 0.8, 0.7, 0.8, 0.9],
            [0.7, 0.9, 0.8, 0.7, 0.8],
        ],
        "k_k": [0.76, 0.82, 0.78],
        "k_a": [0.8, 0.85, 0.8],
        "name_arguments": [
            "Ступінь знайомства",
            "Теоретичний аналіз",
            "Досвід",
            "Література",
            "Інтуїція",
        ],
        "name_research": ["Критерій 1", "Критерій 2", "Критерій 3"],
        "experts_data_table": [[0.8, 0.7, 0.9], [0.9, 0.8, 0.7], [0.7, 0.9, 0.8]],
        "m_i": [0.8, 0.8, 0.8],
        "r_i": [0.8, 0.8, 0.8],
        "l_value": [0.8, 0.8, 0.8],
        "l_value_sum": 2,  # Integer value
        "rank_str": "Критерій 1 та Критерій 2 мають найвищий пріоритет серед всіх аналізованих критеріїв",
    }

    try:
        # Test 1: Create exporter
        print("  ✅ Creating ExpertsExcelExporter...")
        exporter = ExpertsExcelExporter()

        # Test 2: Check text wrapping
        print("  ✅ Checking text wrapping...")

        # Check if wrap_text is enabled in alignment
        from openpyxl.styles import Alignment

        # Create a test cell and apply header style
        from openpyxl import Workbook

        test_wb = Workbook()
        test_ws = test_wb.active
        test_cell = test_ws["A1"]
        exporter.set_header_style(test_cell)

        if hasattr(test_cell.alignment, "wrap_text") and test_cell.alignment.wrap_text:
            print("    ✅ Text wrapping is enabled")
        else:
            print("    ❌ Text wrapping is not enabled")
            return False

        # Test 3: Generate workbook
        print("  ✅ Generating workbook...")
        workbook = exporter.generate_experts_analysis_excel(analysis_data)

        # Test 4: Check conclusion sheet has proper rank_str
        print("  ✅ Checking conclusion sheet...")
        conclusion_ws = workbook["Висновки"]

        if conclusion_ws["A4"].value == analysis_data["rank_str"]:
            print("    ✅ Conclusion has proper rank_str")
        else:
            print("    ❌ Conclusion doesn't have proper rank_str")
            print(f"    Expected: {analysis_data['rank_str']}")
            print(f"    Actual: {conclusion_ws['A4'].value}")
            return False

        # Test 5: Check l_value_sum is integer
        print("  ✅ Checking l_value_sum formatting...")
        data_ws = workbook["Дані експертів"]

        # Find the sum cell (should be in the last column of lambda row)
        lambda_row = None
        for row in range(1, 20):  # Check first 20 rows
            if data_ws[f"A{row}"].value == "λ":
                lambda_row = row
                break

        if lambda_row:
            # Find the sum cell (last column)
            sum_cell = None
            for col in range(2, 10):  # Check columns B to J
                cell_value = data_ws.cell(row=lambda_row, column=col).value
                if cell_value == analysis_data["l_value_sum"]:
                    sum_cell = data_ws.cell(row=lambda_row, column=col)
                    break

            if sum_cell and isinstance(sum_cell.value, int):
                print("    ✅ l_value_sum is formatted as integer")
            else:
                print("    ❌ l_value_sum is not formatted as integer")
                print(f"    Value: {sum_cell.value if sum_cell else 'Not found'}")
                print(f"    Type: {type(sum_cell.value) if sum_cell else 'N/A'}")
                return False
        else:
            print("    ❌ Could not find lambda row")
            return False

        # Test 6: Check text wrapping in all sheets
        print("  ✅ Checking text wrapping in all sheets...")

        sheets_to_check = [
            "Загальна інформація",
            "Компетентність експертів",
            "Дані експертів",
            "Ранжування",
            "Графіки",
            "Висновки",
        ]

        for sheet_name in sheets_to_check:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                # Check a few cells for wrap_text
                cells_to_check = ["A1", "A3", "A4"]
                for cell_ref in cells_to_check:
                    if ws[cell_ref].value:
                        if (
                            hasattr(ws[cell_ref].alignment, "wrap_text")
                            and ws[cell_ref].alignment.wrap_text
                        ):
                            print(f"    ✅ {sheet_name} - {cell_ref} has text wrapping")
                            break
                else:
                    print(
                        f"    ⚠️  {sheet_name} - No text wrapping found in checked cells"
                    )

        # Test 7: Save to bytes
        print("  ✅ Testing save to bytes...")
        excel_bytes = exporter.save_to_bytes()

        if len(excel_bytes) > 1000:
            print(f"    ✅ File size: {len(excel_bytes)} bytes")
        else:
            print(f"    ❌ File too small: {len(excel_bytes)} bytes")
            return False

        print("\n🎉 All final fixes passed!")
        return True

    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_final_fixes()
    if success:
        print("\n✅ All final fixes are working correctly!")
    else:
        print("\n❌ Some final fixes have issues!")
        sys.exit(1)
